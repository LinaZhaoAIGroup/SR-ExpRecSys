#!/usr/bin/env python3
"""Create reproducible, blinded CSV forms for two-expert KG validation.

The script samples unique complete triples from the main KG with a fixed random
seed. It does not alter the KG. Two identical annotation files are created so
that experts can work independently before adjudication.

Because the current triple table has no source-document/provenance columns,
reviewers must locate and record the supporting source passage during review.
Without source evidence, extraction correctness cannot be judged reliably.
"""

from __future__ import annotations

import argparse
import csv
import random
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = SCRIPT_DIR / "02-full-kg" / "BSRF_HEPS_tuple.xlsx"
DEFAULT_SHEET = "Sheet1"
DEFAULT_OUTPUT_DIR = SCRIPT_DIR / "validation" / "round_1"
DEFAULT_SAMPLE_SIZE = 235
DEFAULT_SEED = 20260812
TRIPLE_COLUMNS = (
    "sub_name",
    "sub_type",
    "rel_type",
    "rel_name",
    "obj_name",
    "obj_type",
)
TRIPLE_KEY_COLUMNS = ("sub_name", "rel_name", "obj_name")
WHITESPACE_RE = re.compile(r"\s+")

ANNOTATION_COLUMNS = [
    "sample_id",
    "source_file",
    "source_sheet",
    "source_row",
    "sub_name",
    "sub_type",
    "rel_type",
    "rel_name",
    "obj_name",
    "obj_type",
    "facility_BSRF_or_HEPS",
    "source_document",
    "source_location",
    "source_excerpt",
    "source_support_Y_N_UNCERTAIN",
    "subject_correct_Y_N_UNCERTAIN",
    "relation_correct_Y_N_UNCERTAIN",
    "object_correct_Y_N_UNCERTAIN",
    "entity_types_correct_Y_N_UNCERTAIN",
    "triple_correct_Y_N_UNCERTAIN",
    "error_types_semicolon_separated",
    "action_KEEP_CORRECT_REMOVE_UNCERTAIN",
    "corrected_sub_name",
    "corrected_sub_type",
    "corrected_rel_type",
    "corrected_rel_name",
    "corrected_obj_name",
    "corrected_obj_type",
    "notes",
]

ERROR_TYPE_CODES = [
    "UNSUPPORTED_TRIPLE",
    "WRONG_SUBJECT",
    "WRONG_RELATION",
    "WRONG_OBJECT",
    "WRONG_ENTITY_TYPE",
    "OVERLY_BROAD_OR_NARROW",
    "DUPLICATE",
    "SYNONYM_NOT_MERGED",
    "SELF_LOOP",
    "INCONSISTENT_SCHEMA",
    "MISSING_PROVENANCE",
    "OTHER",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Sample unique triples and create blinded expert-validation forms."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--sample-size", type=int, default=DEFAULT_SAMPLE_SIZE)
    parser.add_argument("--seed", type=int, default=DEFAULT_SEED)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args()


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return WHITESPACE_RE.sub(" ", str(value)).strip()


def load_unique_triples(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    if not path.is_file():
        raise FileNotFoundError(f"Knowledge graph file not found: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet {sheet_name!r} does not exist; available worksheets: "
                f"{workbook.sheetnames}"
            )
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ValueError(f"Worksheet {sheet_name!r} is empty")
        headers = [normalize_text(value).lstrip("\ufeff").lower() for value in header]
        missing = [column for column in TRIPLE_COLUMNS if column not in headers]
        if missing:
            raise ValueError(f"Missing triple fields: {missing}")

        unique: dict[tuple[str, str, str], dict[str, Any]] = {}
        for source_row, row in enumerate(rows, start=2):
            values = list(row) + [None] * max(0, len(headers) - len(row))
            record = {
                name: normalize_text(values[index])
                for index, name in enumerate(headers)
            }
            if not any(record.get(column, "") for column in TRIPLE_COLUMNS):
                continue
            if any(not record.get(column, "") for column in TRIPLE_KEY_COLUMNS):
                continue
            key = tuple(record[column] for column in TRIPLE_KEY_COLUMNS)
            if key not in unique:
                record["source_row"] = source_row
                unique[key] = record
        return list(unique.values())
    finally:
        workbook.close()


def make_annotation_rows(
    sampled: list[dict[str, Any]], source_file: Path, source_sheet: str
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for index, record in enumerate(sampled, start=1):
        output = {column: "" for column in ANNOTATION_COLUMNS}
        output.update(
            {
                "sample_id": f"KG-{index:04d}",
                "source_file": source_file.name,
                "source_sheet": source_sheet,
                "source_row": record["source_row"],
            }
        )
        for column in TRIPLE_COLUMNS:
            output[column] = record[column]
        rows.append(output)
    return rows


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=ANNOTATION_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def write_protocol(path: Path, population: int, sample_size: int, seed: int) -> None:
    error_codes = ", ".join(f"`{code}`" for code in ERROR_TYPE_CODES)
    if population > 1:
        finite_population_correction = (population - sample_size) / (population - 1)
        worst_case_margin = 1.95996398454 * (
            0.25 / sample_size * finite_population_correction
        ) ** 0.5
    else:
        worst_case_margin = 0.0
    path.write_text(
        f"""# Knowledge-Graph Dual-Expert Validation Protocol

## Sampling

- Population: {population:,} complete triples after deduplication by
  `(sub_name, rel_name, obj_name)`.
- Sample size: {sample_size:,} triples.
- Random seed: {seed}.
- Method: simple random sampling without replacement. Under the conservative
  proportion of 0.5, the approximate finite-population 95% margin of error is
  +/-{worst_case_margin:.1%}. Final estimates also report Wilson 95% intervals.
- The released table lacks source documents and paragraph identifiers, so this
  sample can estimate triple precision but cannot establish recall.

## Independent Review

Two experts should complete `expert_1_annotations.csv` and
`expert_2_annotations.csv` independently. Before assigning a correctness label,
record `facility_BSRF_or_HEPS`, `source_document`, `source_location`, and
`source_excerpt`. If the source cannot be verified, mark both source support and
triple correctness as `UNCERTAIN`.

Use `Y` for a triple whose subject, relation, object, and entity types are all
correct and source-supported. Use `N` when any core component is wrong. Use
`UNCERTAIN` when the evidence is insufficient. The action field accepts only
`KEEP`, `CORRECT`, `REMOVE`, or `UNCERTAIN`; `CORRECT` requires at least one
corrected field. Error types are semicolon-separated controlled codes:
{error_codes}.

## Agreement and Adjudication

After both reviews, run `calculate_kg_validation_metrics.py`. It reports strict
triple precision, component-level precision, Wilson intervals, agreement, and
Cohen's kappa, and writes a disagreement table. After adjudication, complete
`adjudication_template.csv` and rerun the script with `--adjudicated`.

Full-map corrections, removals, synonym merges, and relation normalizations
must be recorded one by one in `kg_cleaning_log_template.csv`; sample actions
alone cannot support full-map counts.

## Recall Limitations

Recall requires an independent random sample of source passages or other source
units. Experts must enumerate all gold-standard triples without seeing system
outputs. The current triple table has no paragraph-level provenance, so this
workflow alone cannot provide a reliable recall estimate.
""",
        encoding="utf-8",
    )


def write_recall_templates(output_dir: Path) -> None:
    source_unit_columns = [
        "source_unit_id",
        "included_in_random_sample_Y_N",
        "facility_BSRF_or_HEPS",
        "source_document",
        "source_location",
        "source_text",
        "expert_1_completed_Y_N",
        "expert_2_completed_Y_N",
        "adjudicated_Y_N",
        "notes",
    ]
    with (output_dir / "recall_source_units.csv").open(
        "w", encoding="utf-8-sig", newline=""
    ) as handle:
        csv.writer(handle).writerow(source_unit_columns)

    columns = [
        "source_unit_id",
        "gold_triple_id",
        "gold_sub_name",
        "gold_sub_type",
        "gold_rel_name",
        "gold_rel_type",
        "gold_obj_name",
        "gold_obj_type",
        "matched_by_system_Y_N_UNCERTAIN",
        "matched_system_sub_name",
        "matched_system_rel_name",
        "matched_system_obj_name",
        "notes",
    ]
    with (output_dir / "recall_gold_triples.csv").open(
        "w", encoding="utf-8-sig", newline=""
    ) as handle:
        csv.writer(handle).writerow(columns)


def main() -> None:
    args = parse_args()
    if args.sample_size <= 0:
        raise ValueError("sample-size Must be a positive integer")

    source_file = args.input.resolve()
    population = load_unique_triples(source_file, args.sheet)
    if args.sample_size > len(population):
        raise ValueError(
            f"sample size{args.sample_size} is greater than the total number of unique complete triples{len(population)}"
        )

    rng = random.Random(args.seed)
    selected_indexes = sorted(rng.sample(range(len(population)), args.sample_size))
    sampled = [population[index] for index in selected_indexes]
    annotation_rows = make_annotation_rows(sampled, source_file, args.sheet)

    output_dir = args.output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    write_csv(output_dir / "expert_1_annotations.csv", annotation_rows)
    write_csv(output_dir / "expert_2_annotations.csv", annotation_rows)
    write_protocol(
        output_dir / "annotation_protocol.md",
        population=len(population),
        sample_size=args.sample_size,
        seed=args.seed,
    )
    write_recall_templates(output_dir)
    print(f"Verification materials have been generated:{output_dir}")


if __name__ == "__main__":
    main()
