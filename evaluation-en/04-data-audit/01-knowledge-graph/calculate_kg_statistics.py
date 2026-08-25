#!/usr/bin/env python3
"""Audit the size and structural consistency of the BSRF-HEPS KG.

The script is read-only with respect to the source CSV/XLSX files. It reports
raw rows, complete triples, exact duplicate triples, unique nodes/edges,
entity-type and relation-label distributions, and several quality flags. It
also includes the size of the K-V source corpus when that file is available.

Important: these structural statistics do not measure extraction precision or
recall. Use prepare_kg_validation_sample.py and
calculate_kg_validation_metrics.py for expert validation.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_MAIN_KG = SCRIPT_DIR / "02-full-kg" / "BSRF_HEPS_tuple.xlsx"
DEFAULT_MAIN_SHEET = "Sheet1"
DEFAULT_EXPERIMENTAL_KG = (
    SCRIPT_DIR / "01-experimental-technique-kg" / "experimental_technique_tuple.csv"
)
DEFAULT_CORPUS = (
    SCRIPT_DIR.parent / "00-corpus-scale" / "BSRF-HEPS-KnowledgeBase-k-v.xlsx"
)
DEFAULT_CORPUS_SHEET = "Sheet1"
DEFAULT_OUTPUT_DIR = SCRIPT_DIR / "outputs" / "kg_statistics"

TRIPLE_COLUMNS = (
    "sub_name",
    "sub_type",
    "rel_type",
    "rel_name",
    "obj_name",
    "obj_type",
)
TRIPLE_KEY_COLUMNS = ("sub_name", "rel_name", "obj_name")
PLACEHOLDER_VALUES = {"[Picture placeholder]", "——"}
WHITESPACE_RE = re.compile(r"\s+")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Statistical knowledge graph scale and output relationships/Entity category distribution and quality audit details."
    )
    parser.add_argument("--main-kg", type=Path, default=DEFAULT_MAIN_KG)
    parser.add_argument("--main-sheet", default=DEFAULT_MAIN_SHEET)
    parser.add_argument(
        "--experimental-kg", type=Path, default=DEFAULT_EXPERIMENTAL_KG
    )
    parser.add_argument("--skip-experimental", action="store_true")
    parser.add_argument("--corpus", type=Path, default=DEFAULT_CORPUS)
    parser.add_argument("--corpus-sheet", default=DEFAULT_CORPUS_SHEET)
    parser.add_argument("--skip-corpus", action="store_true")
    parser.add_argument("--top-relations", type=int, default=15)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args()


def normalize_text(value: Any) -> str:
    """Normalize cell text without applying semantic synonym merging."""
    if value is None:
        return ""
    return WHITESPACE_RE.sub(" ", str(value)).strip()


def normalized_header(value: Any) -> str:
    return normalize_text(value).lstrip("\ufeff").lower()


def load_xlsx_table(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    if not path.is_file():
        raise FileNotFoundError(f"not foundExcel File:{path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"worksheet{sheet_name!r} Does not exist; available worksheets:{workbook.sheetnames}"
            )
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ValueError(f"worksheet{sheet_name!r} is empty")
        all_headers = [normalized_header(value) for value in header]
        header_positions = [
            (index, header_name)
            for index, header_name in enumerate(all_headers)
            if header_name
        ]
        if not header_positions:
            raise ValueError(f"worksheet{sheet_name!r} No valid header")

        headers = [header_name for _, header_name in header_positions]
        duplicate_headers = sorted(
            {header_name for header_name in headers if headers.count(header_name) > 1}
        )
        if duplicate_headers:
            raise ValueError(f"There are duplicate columns in the header:{duplicate_headers}")

        records: list[dict[str, Any]] = []
        for excel_row, row in enumerate(rows, start=2):
            record = {
                header_name: normalize_text(row[index] if index < len(row) else None)
                for index, header_name in header_positions
            }
            if not any(record.values()):
                continue
            record["source_row"] = excel_row
            records.append(record)
        return records
    finally:
        workbook.close()


def load_csv_table(path: Path) -> list[dict[str, Any]]:
    if not path.is_file():
        raise FileNotFoundError(f"not foundCSV File:{path}")

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise ValueError(f"CSV The file has no header:{path}")
        header_map = {name: normalized_header(name) for name in reader.fieldnames}
        records: list[dict[str, Any]] = []
        for source_row, row in enumerate(reader, start=2):
            record = {
                header_map[name]: normalize_text(row.get(name))
                for name in reader.fieldnames
            }
            if not any(record.values()):
                continue
            record["source_row"] = source_row
            records.append(record)
        return records


def load_table(path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return load_csv_table(path)
    if suffix in {".xlsx", ".xlsm"}:
        if sheet_name is None:
            raise ValueError("readExcel The worksheet name must be specified when")
        return load_xlsx_table(path, sheet_name)
    raise ValueError(f"Unsupported file types:{path.suffix}")


def validate_triple_columns(records: list[dict[str, Any]], source: Path) -> None:
    if not records:
        raise ValueError(f"No data was read:{source}")
    missing = [column for column in TRIPLE_COLUMNS if column not in records[0]]
    if missing:
        raise ValueError(f"{source.name} Missing triplet field:{missing}")


def triple_key(record: dict[str, Any]) -> tuple[str, str, str]:
    return tuple(record[column] for column in TRIPLE_KEY_COLUMNS)  # type: ignore[return-value]


def sorted_count_rows(counter: Counter[str], label_column: str) -> list[dict[str, Any]]:
    return [
        {label_column: label, "count": count}
        for label, count in sorted(counter.items(), key=lambda item: (-item[1], item[0]))
    ]


def audit_triples(
    records: list[dict[str, Any]], dataset_name: str, source: Path, sheet: str | None
) -> dict[str, Any]:
    validate_triple_columns(records, source)

    incomplete = [
        record
        for record in records
        if any(not record[column] for column in TRIPLE_KEY_COLUMNS)
    ]
    complete = [
        record
        for record in records
        if all(record[column] for column in TRIPLE_KEY_COLUMNS)
    ]

    rows_by_key: defaultdict[tuple[str, str, str], list[int]] = defaultdict(list)
    first_record_by_key: dict[tuple[str, str, str], dict[str, Any]] = {}
    for record in complete:
        key = triple_key(record)
        rows_by_key[key].append(record["source_row"])
        first_record_by_key.setdefault(key, record)
    unique_records = list(first_record_by_key.values())

    duplicate_groups = []
    for key, source_rows in rows_by_key.items():
        if len(source_rows) > 1:
            duplicate_groups.append(
                {
                    "sub_name": key[0],
                    "rel_name": key[1],
                    "obj_name": key[2],
                    "occurrences": len(source_rows),
                    "extra_duplicate_rows": len(source_rows) - 1,
                    "source_rows": ";".join(str(value) for value in source_rows),
                }
            )

    nodes: set[str] = set()
    typed_nodes: set[tuple[str, str]] = set()
    class_to_nodes: defaultdict[str, set[str]] = defaultdict(set)
    types_by_node: defaultdict[str, set[str]] = defaultdict(set)
    relation_name_counts: Counter[str] = Counter()
    relation_type_counts: Counter[str] = Counter()
    relation_pair_counts: Counter[tuple[str, str]] = Counter()
    self_loops: list[dict[str, Any]] = []
    relation_mismatches: list[dict[str, Any]] = []
    placeholder_rows: list[dict[str, Any]] = []

    for record in unique_records:
        for name_column, type_column in (
            ("sub_name", "sub_type"),
            ("obj_name", "obj_type"),
        ):
            name = record[name_column]
            entity_type = record[type_column]
            if name:
                nodes.add(name)
            if name and entity_type:
                typed_nodes.add((name, entity_type))
                class_to_nodes[entity_type].add(name)
                types_by_node[name].add(entity_type)

        relation_name_counts[record["rel_name"]] += 1
        relation_type_counts[record["rel_type"]] += 1
        relation_pair_counts[(record["rel_type"], record["rel_name"])] += 1

        if record["sub_name"] == record["obj_name"]:
            self_loops.append(record)
        if record["rel_type"] != record["rel_name"]:
            relation_mismatches.append(record)
        if record["sub_name"] in PLACEHOLDER_VALUES or record["obj_name"] in PLACEHOLDER_VALUES:
            placeholder_rows.append(record)

    type_conflicts = [
        {
            "entity_name": name,
            "type_count": len(entity_types),
            "entity_types": ";".join(sorted(entity_types)),
        }
        for name, entity_types in sorted(types_by_node.items())
        if len(entity_types) > 1
    ]

    entity_class_rows = [
        {"entity_type": entity_type, "unique_node_count": len(class_to_nodes[entity_type])}
        for entity_type in sorted(
            class_to_nodes, key=lambda value: (-len(class_to_nodes[value]), value)
        )
    ]
    relation_name_rows = sorted_count_rows(relation_name_counts, "relation_name")
    relation_type_rows = sorted_count_rows(relation_type_counts, "relation_type")
    relation_mapping_rows = [
        {
            "raw_rel_type": relation_type,
            "raw_rel_name": relation_name,
            "unique_triple_count": count,
            "canonical_relation": "",
            "decision_keep_merge_remove": "",
            "notes": "",
        }
        for (relation_type, relation_name), count in sorted(
            relation_pair_counts.items(), key=lambda item: (-item[1], item[0])
        )
    ]
    entity_mapping_rows = [
        {
            "raw_entity_type": row["entity_type"],
            "unique_node_count": row["unique_node_count"],
            "canonical_entity_class": "",
            "decision_keep_merge_remove": "",
            "notes": "",
        }
        for row in entity_class_rows
    ]

    quality_flags: list[dict[str, Any]] = []
    for record in incomplete:
        quality_flags.append(
            quality_row("incomplete_triple", record, "subject/relation/object At least one item is empty")
        )
    for group in duplicate_groups:
        quality_flags.append(
            {
                "issue": "exact_duplicate_group",
                "source_row": group["source_rows"],
                "sub_name": group["sub_name"],
                "rel_type": "",
                "rel_name": group["rel_name"],
                "obj_name": group["obj_name"],
                "detail": f"appear{group['occurrences']} times",
            }
        )
    for record in self_loops:
        quality_flags.append(quality_row("self_loop_candidate", record, "subject withobject Same"))
    for record in relation_mismatches:
        quality_flags.append(
            quality_row("relation_label_mismatch", record, "rel_type withrel_name Not the same")
        )
    for record in placeholder_rows:
        quality_flags.append(
            quality_row("placeholder_entity", record, "subject orobject for placeholder text")
        )
    for conflict in type_conflicts:
        quality_flags.append(
            {
                "issue": "entity_type_conflict",
                "source_row": "",
                "sub_name": conflict["entity_name"],
                "rel_type": "",
                "rel_name": "",
                "obj_name": "",
                "detail": conflict["entity_types"],
            }
        )

    summary = {
        "dataset": dataset_name,
        "source_file": source.name,
        "sheet": sheet,
        "raw_nonblank_rows": len(records),
        "complete_triple_rows": len(complete),
        "incomplete_triple_rows": len(incomplete),
        "unique_directed_triples_edges": len(unique_records),
        "exact_duplicate_groups": len(duplicate_groups),
        "extra_exact_duplicate_rows": sum(
            row["extra_duplicate_rows"] for row in duplicate_groups
        ),
        "unique_node_names": len(nodes),
        "unique_typed_node_assignments": len(typed_nodes),
        "observed_entity_type_labels": len(class_to_nodes),
        "entity_names_with_multiple_types": len(type_conflicts),
        "observed_relation_name_labels": len(relation_name_counts),
        "observed_relation_type_labels": len(relation_type_counts),
        "relation_type_name_mismatch_edges": len(relation_mismatches),
        "self_loop_candidates": len(self_loops),
        "placeholder_entity_edges": len(placeholder_rows),
    }

    return {
        "summary": summary,
        "entity_class_distribution": entity_class_rows,
        "relation_name_distribution": relation_name_rows,
        "relation_type_distribution": relation_type_rows,
        "duplicate_groups": duplicate_groups,
        "entity_type_conflicts": type_conflicts,
        "quality_flags": quality_flags,
        "relation_mapping_template": relation_mapping_rows,
        "entity_type_mapping_template": entity_mapping_rows,
    }


def quality_row(issue: str, record: dict[str, Any], detail: str) -> dict[str, Any]:
    return {
        "issue": issue,
        "source_row": record.get("source_row", ""),
        "sub_name": record.get("sub_name", ""),
        "rel_type": record.get("rel_type", ""),
        "rel_name": record.get("rel_name", ""),
        "obj_name": record.get("obj_name", ""),
        "detail": detail,
    }


def audit_corpus(path: Path, sheet_name: str) -> dict[str, Any]:
    records = load_xlsx_table(path, sheet_name)
    if not records:
        raise ValueError(f"The corpus is empty:{path}")
    missing = [column for column in ("k", "v") if column not in records[0]]
    if missing:
        raise ValueError(f"The corpus is missing a field:{missing}")

    complete = [record for record in records if record["k"] and record["v"]]
    substantive = [
        record for record in complete if record["v"] not in PLACEHOLDER_VALUES
    ]
    unique_pairs = {(record["k"], record["v"]) for record in complete}
    return {
        "source_file": path.name,
        "sheet": sheet_name,
        "confirmed_as_complete_kg_extraction_source": False,
        "confirmation_note": (
            "The KG table contains no source provenance. Confirm that this K-V file is the "
            "complete extraction source before reporting it as the source-corpus size."
        ),
        "complete_kv_records": len(complete),
        "unique_kv_pairs": len(unique_pairs),
        "substantive_kv_records": len(substantive),
        "placeholder_records": len(complete) - len(substantive),
        "total_kv_characters": sum(
            len(record["k"]) + len(record["v"]) for record in complete
        ),
        "substantive_kv_characters": sum(
            len(record["k"]) + len(record["v"]) for record in substantive
        ),
    }


def write_csv(path: Path, rows: Iterable[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_dataset_outputs(
    output_dir: Path, prefix: str, report: dict[str, Any]
) -> None:
    write_csv(
        output_dir / f"{prefix}_entity_class_distribution.csv",
        report["entity_class_distribution"],
        ["entity_type", "unique_node_count"],
    )
    write_csv(
        output_dir / f"{prefix}_relation_name_distribution.csv",
        report["relation_name_distribution"],
        ["relation_name", "count"],
    )
    write_csv(
        output_dir / f"{prefix}_relation_type_distribution.csv",
        report["relation_type_distribution"],
        ["relation_type", "count"],
    )
    write_csv(
        output_dir / f"{prefix}_quality_flags.csv",
        report["quality_flags"],
        [
            "issue",
            "source_row",
            "sub_name",
            "rel_type",
            "rel_name",
            "obj_name",
            "detail",
        ],
    )
    write_csv(
        output_dir / f"{prefix}_relation_mapping_template.csv",
        report["relation_mapping_template"],
        [
            "raw_rel_type",
            "raw_rel_name",
            "unique_triple_count",
            "canonical_relation",
            "decision_keep_merge_remove",
            "notes",
        ],
    )
    write_csv(
        output_dir / f"{prefix}_entity_type_mapping_template.csv",
        report["entity_type_mapping_template"],
        [
            "raw_entity_type",
            "unique_node_count",
            "canonical_entity_class",
            "decision_keep_merge_remove",
            "notes",
        ],
    )


def markdown_table(headers: list[str], rows: list[list[Any]]) -> list[str]:
    lines = [
        "| " + " | ".join(headers) + " |",
        "|" + "|".join("---" for _ in headers) + "|",
    ]
    for row in rows:
        lines.append("| " + " | ".join(str(value) for value in row) + " |")
    return lines


def build_markdown(
    reports: dict[str, dict[str, Any]],
    corpus: dict[str, Any] | None,
    top_relations: int,
) -> str:
    lines = [
        "# BSRF-HEPS Knowledge-Graph Scale and Structure Audit",
        "",
        "## Counting Rules",
        "",
        "- The main graph is read from `Sheet1` of `BSRF_HEPS_tuple.xlsx`.",
        "- Unique directed edges are counted after whitespace normalization using `(sub_name, rel_name, obj_name)`.",
        "- Unique nodes are counted by normalized entity name, consistent with the graph loader's name-based deduplication.",
        "- Typed nodes are unique `(entity name, entity type)` pairs; one entity name may therefore contribute multiple typed nodes.",
        "- Relation names and entity types use the original table labels; no semantic merging is applied.",
        "- This report is a structural audit and does not represent extraction precision, recall, or expert agreement.",
        "",
        "## Overall statistics",
        "",
    ]

    table_rows: list[list[Any]] = []
    for report in reports.values():
        summary = report["summary"]
        table_rows.append(
            [
                summary["dataset"],
                summary["raw_nonblank_rows"],
                summary["complete_triple_rows"],
                summary["unique_directed_triples_edges"],
                summary["unique_node_names"],
                summary["observed_entity_type_labels"],
                summary["observed_relation_name_labels"],
            ]
        )
    lines.extend(
        markdown_table(
            [
                "Dataset",
                "Nonblank rows",
                "Complete triples",
                "Unique directed edges",
                "Unique nodes",
                "Entity-type labels",
                "Relation-name labels",
            ],
            table_rows,
        )
    )

    if corpus is not None:
        lines.extend(
            [
                "",
                "## Candidate Source-Corpus Size",
                "",
            ]
        )
        lines.extend(
            markdown_table(
                ["Metric", "Value"],
                [
                    ["Complete K-V records", corpus["complete_kv_records"]],
                    ["Unique K-V pairs", corpus["unique_kv_pairs"]],
                    ["Substantive K-V records", corpus["substantive_kv_records"]],
                    ["Placeholder records", corpus["placeholder_records"]],
                    ["Total K-V characters", f"{corpus['total_kv_characters']:,}"],
                    [
                        "Substantive K-V characters",
                        f"{corpus['substantive_kv_characters']:,}",
                    ],
                ],
            )
        )
        lines.extend(
            [
                "",
                "Note: the graph table has no source-document or paragraph identifiers, so the script cannot verify whether the K-V file is the complete extraction source. Do not interpret these values as graph source-corpus statistics without independent confirmation.",
            ]
        )

    for key, report in reports.items():
        summary = report["summary"]
        lines.extend(
            [
                "",
                f"## {summary['dataset']}",
                "",
                f"Quality audit: {summary['extra_exact_duplicate_rows']} extra exact duplicate rows; "
                f"{summary['entity_names_with_multiple_types']} entity names with multiple types; "
                f"{summary['self_loop_candidates']} self-loop candidates; and "
                f"{summary['relation_type_name_mismatch_edges']} unique edges with differing `rel_type` and `rel_name` values.",
                "",
                "### Entity type distribution",
                "",
            ]
        )
        lines.extend(
            markdown_table(
                ["Raw entity type", "Unique nodes"],
                [
                    [row["entity_type"] or "[empty]", row["unique_node_count"]]
                    for row in report["entity_class_distribution"]
                ],
            )
        )
        lines.extend(["", f"### Top {top_relations} Relation Names", ""])
        lines.extend(
            markdown_table(
                ["Raw relation name", "Unique edges"],
                [
                    [row["relation_name"] or "[empty]", row["count"]]
                    for row in report["relation_name_distribution"][:top_relations]
                ],
            )
        )
        remaining = report["relation_name_distribution"][top_relations:]
        if remaining:
            lines.append("")
            lines.append(
                f"The remaining {len(remaining)} relation names account for "
                f"{sum(row['count'] for row in remaining)} unique edges. See the CSV output for the complete distribution."
            )

    lines.extend(
        [
            "",
            "## Interpretation Limits",
            "",
            "The raw table contains more entity-type and relation labels than the core ontology categories discussed in the manuscript, as well as duplicate records, self-loop candidates, and names assigned to multiple types. Domain experts should complete label mapping and exception adjudication before final category counts are reported. Historical synonym merges, corrections, and removals cannot be reconstructed from the final triple table alone.",
            "",
        ]
    )
    return "\n".join(lines)


def main() -> None:
    args = parse_args()
    output_dir = args.output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    reports: dict[str, dict[str, Any]] = {}
    main_records = load_table(args.main_kg.resolve(), args.main_sheet)
    reports["main"] = audit_triples(
        main_records,
        dataset_name="Main knowledge graph",
        source=args.main_kg.resolve(),
        sheet=args.main_sheet,
    )

    if not args.skip_experimental and args.experimental_kg.is_file():
        experimental_records = load_table(args.experimental_kg.resolve())
        reports["experimental"] = audit_triples(
            experimental_records,
            dataset_name="Experimental technology special map",
            source=args.experimental_kg.resolve(),
            sheet=None,
        )

    corpus = None
    if not args.skip_corpus and args.corpus.is_file():
        corpus = audit_corpus(args.corpus.resolve(), args.corpus_sheet)

    for prefix, report in reports.items():
        write_dataset_outputs(output_dir, prefix, report)

    complete_report = {
        "definitions": {
            "edge_identity": list(TRIPLE_KEY_COLUMNS),
            "node_identity": "normalized entity name",
            "typed_node_identity": ["entity name", "entity type"],
            "semantic_synonym_merging_applied": False,
        },
        "source_corpus": corpus,
        "datasets": reports,
    }
    (output_dir / "kg_statistics_report.json").write_text(
        json.dumps(complete_report, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (output_dir / "KG_statistics_summary.md").write_text(
        build_markdown(reports, corpus, args.top_relations), encoding="utf-8"
    )

    print(f"Statistics completed:{output_dir}")


if __name__ == "__main__":
    main()
