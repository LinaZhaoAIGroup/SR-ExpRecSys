#!/usr/bin/env python3
"""Measure the size and composition of the BSRF-HEPS K-V corpus."""

from __future__ import annotations

import argparse
import json
import math
import re
from collections import Counter, defaultdict
from pathlib import Path
from statistics import mean, median
from typing import Any, Iterable

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = SCRIPT_DIR / "BSRF-HEPS-KnowledgeBase-k-v.xlsx"
DEFAULT_SHEET = "Sheet1"
DEFAULT_JSON_OUTPUT = SCRIPT_DIR / "corpus_scale_report.json"
DEFAULT_DESCRIPTION_OUTPUT = SCRIPT_DIR / "corpus_scale_description.md"
PLACEHOLDER_VALUES = {"[Picture placeholder]", "——"}

WHITESPACE_RE = re.compile(r"\s")
CHINESE_CHAR_RE = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff]")
ENGLISH_WORD_RE = re.compile(r"[A-Za-z]+(?:[-'][A-Za-z]+)*")
NUMBER_RE = re.compile(r"\d+(?:\.\d+)?")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Measure the K-V corpus and generate a Markdown summary."
    )
    parser.add_argument(
        "input_file",
        nargs="?",
        type=Path,
        default=DEFAULT_INPUT,
        help=f"Input Excel workbook (default: {DEFAULT_INPUT.name})",
    )
    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET,
        help=f"Worksheet name (default: {DEFAULT_SHEET})",
    )
    parser.add_argument(
        "--json-output",
        type=Path,
        default=DEFAULT_JSON_OUTPUT,
        help=f"JSON report path (default: {DEFAULT_JSON_OUTPUT.name})",
    )
    parser.add_argument(
        "--description-output",
        type=Path,
        default=DEFAULT_DESCRIPTION_OUTPUT,
        help=f"Markdown summary path (default: {DEFAULT_DESCRIPTION_OUTPUT.name})",
    )
    parser.add_argument(
        "--no-save",
        action="store_true",
        help="Print the report without writing JSON or Markdown files",
    )
    return parser.parse_args()


def as_text(value: Any) -> str:
    """Convert an Excel value to stripped text."""
    return "" if value is None else str(value).strip()


def safe_ratio(numerator: int, denominator: int) -> float:
    return round(numerator / denominator, 6) if denominator else 0.0


def percentile(numbers: list[int], proportion: float) -> float:
    """Quantiles are calculated using linear interpolation."""
    if not numbers:
        return 0.0
    ordered = sorted(numbers)
    position = (len(ordered) - 1) * proportion
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return float(ordered[lower])
    weight = position - lower
    return ordered[lower] * (1 - weight) + ordered[upper] * weight


def rounded(value: float) -> int | float:
    rounded_value = round(value, 2)
    return int(rounded_value) if rounded_value.is_integer() else rounded_value


def summarize_lengths(texts: list[str]) -> dict[str, int | float]:
    lengths = [len(text) for text in texts]
    if not lengths:
        return {
            "total": 0,
            "mean": 0,
            "median": 0,
            "min": 0,
            "p25": 0,
            "p75": 0,
            "p90": 0,
            "p95": 0,
            "max": 0,
        }
    return {
        "total": sum(lengths),
        "mean": rounded(mean(lengths)),
        "median": rounded(float(median(lengths))),
        "min": min(lengths),
        "p25": rounded(percentile(lengths, 0.25)),
        "p75": rounded(percentile(lengths, 0.75)),
        "p90": rounded(percentile(lengths, 0.90)),
        "p95": rounded(percentile(lengths, 0.95)),
        "max": max(lengths),
    }


def summarize_text(texts: list[str]) -> dict[str, Any]:
    total_characters = sum(len(text) for text in texts)
    whitespace_characters = sum(len(WHITESPACE_RE.findall(text)) for text in texts)
    chinese_characters = sum(len(CHINESE_CHAR_RE.findall(text)) for text in texts)
    english_words = sum(len(ENGLISH_WORD_RE.findall(text)) for text in texts)
    number_sequences = sum(len(NUMBER_RE.findall(text)) for text in texts)

    return {
        "records": len(texts),
        "length": summarize_lengths(texts),
        "characters": {
            "total": total_characters,
            "without_whitespace": total_characters - whitespace_characters,
            "whitespace": whitespace_characters,
            "chinese": chinese_characters,
            "english_word_units": english_words,
            "number_sequences": number_sequences,
        },
    }


def summarize_duplicates(items: Iterable[Any]) -> dict[str, int | float]:
    values = list(items)
    counts = Counter(values)
    repeated_counts = [count for count in counts.values() if count > 1]
    extra_duplicates = sum(count - 1 for count in repeated_counts)
    return {
        "unique_count": len(counts),
        "repeated_groups": len(repeated_counts),
        "rows_in_repeated_groups": sum(repeated_counts),
        "extra_duplicate_rows": extra_duplicates,
        "duplicate_rate": safe_ratio(extra_duplicates, len(values)),
    }


def try_token_count(texts: list[str]) -> dict[str, Any]:
    """Optionally usecl100k_base Estimatetoken number, no additional dependencies are required to be installed."""
    try:
        import tiktoken  # type: ignore
    except ImportError:
        return {
            "available": False,
            "encoding": None,
            "total": None,
            "mean_per_record": None,
            "note": "Not installedtiktoken, Approximation not calculatedtoken Count.",
        }

    encoding_name = "cl100k_base"
    try:
        encoder = tiktoken.get_encoding(encoding_name)
    except Exception:
        return {
            "available": False,
            "encoding": encoding_name,
            "total": None,
            "mean_per_record": None,
            "note": "The tokenizer data are unavailable locally; token count was omitted.",
        }
    token_lengths = [len(encoder.encode(text)) for text in texts]
    total = sum(token_lengths)
    return {
        "available": True,
        "encoding": encoding_name,
        "total": total,
        "mean_per_record": rounded(mean(token_lengths)) if token_lengths else 0,
        "note": (
            "thetoken The number is onlycl100k_base Reproducible estimate of caliber,"
            "Not representative of other modelstokenizer result."
        ),
    }


def load_records(
    input_file: Path, sheet_name: str
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    if not input_file.is_file():
        raise FileNotFoundError(f"Input file not found:{input_file}")

    workbook = load_workbook(input_file, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"worksheet{sheet_name!r} Does not exist, available worksheets:{workbook.sheetnames}"
            )

        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ValueError(f"worksheet{sheet_name!r} is empty")

        headers = [as_text(value).lower() for value in header]
        try:
            key_index = headers.index("k")
            value_index = headers.index("v")
        except ValueError as exc:
            raise ValueError(
                f"The first row of the worksheet must containk andv Column, the actual header is:{headers}"
            ) from exc

        records: list[dict[str, Any]] = []
        blank_rows = 0
        for excel_row, row in enumerate(rows, start=2):
            key = as_text(row[key_index] if key_index < len(row) else None)
            value = as_text(row[value_index] if value_index < len(row) else None)
            if not key and not value:
                blank_rows += 1
                continue
            records.append(
                {
                    "excel_row": excel_row,
                    "key": key,
                    "value": value,
                    "complete": bool(key and value),
                    "placeholder": value in PLACEHOLDER_VALUES,
                }
            )

        metadata = {
            "file": input_file.name,
            "file_size_bytes": input_file.stat().st_size,
            "file_size_kib": round(input_file.stat().st_size / 1024, 2),
            "sheet": sheet_name,
            "worksheet_max_row": worksheet.max_row,
            "worksheet_max_column": worksheet.max_column,
            "blank_rows_within_used_range": blank_rows,
        }
        return metadata, records
    finally:
        workbook.close()


def calculate_scale(input_file: Path, sheet_name: str) -> dict[str, Any]:
    metadata, records = load_records(input_file, sheet_name)
    complete_records = [record for record in records if record["complete"]]
    substantive_records = [
        record for record in complete_records if not record["placeholder"]
    ]

    if not complete_records:
        raise ValueError("not foundk andv Complete corpus records that are both non-empty")

    keys = [record["key"] for record in complete_records]
    values = [record["value"] for record in complete_records]
    pairs = [(record["key"], record["value"]) for record in complete_records]
    # K-V The total number of characters isK The number of characters is the same asV The statistical caliber of adding the number of characters directly.
    combined = [key + value for key, value in pairs]

    substantive_keys = [record["key"] for record in substantive_records]
    substantive_values = [record["value"] for record in substantive_records]
    substantive_pairs = [
        (record["key"], record["value"]) for record in substantive_records
    ]
    substantive_combined = [key + value for key, value in substantive_pairs]

    values_by_key: defaultdict[str, set[str]] = defaultdict(set)
    for key, value in pairs:
        values_by_key[key].add(value)
    one_to_many_keys = {
        key: len(distinct_values)
        for key, distinct_values in values_by_key.items()
        if len(distinct_values) > 1
    }

    placeholder_counts = Counter(
        record["value"] for record in complete_records if record["placeholder"]
    )
    missing_key_rows = [record["excel_row"] for record in records if not record["key"]]
    missing_value_rows = [
        record["excel_row"] for record in records if not record["value"]
    ]

    return {
        "metadata": metadata,
        "record_scale": {
            "nonempty_data_rows": len(records),
            "complete_kv_records": len(complete_records),
            "incomplete_records": len(records) - len(complete_records),
            "missing_key_rows": missing_key_rows,
            "missing_value_rows": missing_value_rows,
            "substantive_kv_records": len(substantive_records),
            "substantive_record_rate": safe_ratio(
                len(substantive_records), len(complete_records)
            ),
            "placeholder_records": len(complete_records) - len(substantive_records),
            "placeholder_rate": safe_ratio(
                len(complete_records) - len(substantive_records),
                len(complete_records),
            ),
            "placeholder_counts": dict(placeholder_counts),
        },
        "deduplication_scale": {
            "keys": summarize_duplicates(keys),
            "values": summarize_duplicates(values),
            "kv_pairs": summarize_duplicates(pairs),
            "one_to_many_key_count": len(one_to_many_keys),
            "one_to_many_key_rate": safe_ratio(len(one_to_many_keys), len(set(keys))),
            "substantive_unique_keys": len(set(substantive_keys)),
            "substantive_unique_values": len(set(substantive_values)),
            "substantive_unique_kv_pairs": len(set(substantive_pairs)),
        },
        "text_scale_all_records": {
            "key": summarize_text(keys),
            "value": summarize_text(values),
            "key_value": summarize_text(combined),
            "approximate_tokens": try_token_count(combined),
        },
        "text_scale_substantive_records": {
            "key": summarize_text(substantive_keys),
            "value": summarize_text(substantive_values),
            "key_value": summarize_text(substantive_combined),
            "approximate_tokens": try_token_count(substantive_combined),
        },
    }


def build_description(report: dict[str, Any]) -> str:
    records = report["record_scale"]
    dedup = report["deduplication_scale"]
    all_text = report["text_scale_all_records"]
    substantive_text = report["text_scale_substantive_records"]

    kv = all_text["key_value"]
    key = all_text["key"]
    value = all_text["value"]
    substantive_kv = substantive_text["key_value"]

    paragraphs = [
        (
            "The BSRF-HEPS key-value corpus contains "
            f"{records['complete_kv_records']:,} complete K-V records, "
            f"{dedup['keys']['unique_count']:,} unique keys, "
            f"{dedup['values']['unique_count']:,} unique values, and "
            f"{dedup['kv_pairs']['unique_count']:,} unique K-V pairs. "
            f"There are {dedup['kv_pairs']['extra_duplicate_rows']:,} exact duplicate "
            f"rows ({dedup['kv_pairs']['duplicate_rate']:.2%}) and "
            f"{dedup['one_to_many_key_count']:,} keys associated with multiple values."
        ),
        (
            f"The key column contains {key['characters']['total']:,} characters and "
            f"the value column contains {value['characters']['total']:,} characters. "
            f"Together they contain {kv['characters']['total']:,} characters, or "
            f"{kv['characters']['without_whitespace']:,} after whitespace removal. "
            f"A K-V record contains a mean of {kv['length']['mean']} characters and a "
            f"median of {kv['length']['median']}; the range is "
            f"{kv['length']['min']}-{kv['length']['max']}, with P90 and P95 values of "
            f"{kv['length']['p90']} and {kv['length']['p95']}."
        ),
        (
            f"The corpus contains {kv['characters']['chinese']:,} Chinese characters, "
            f"{kv['characters']['english_word_units']:,} English word units, and "
            f"{kv['characters']['number_sequences']:,} numeric sequences. After "
            'excluding placeholder values such as "[Picture placeholder]" and "——", '
            f"{records['substantive_kv_records']:,} substantive records remain "
            f"({records['substantive_record_rate']:.2%} of complete records), containing "
            f"{substantive_kv['characters']['total']:,} K-V characters."
        ),
    ]
    return "\n\n".join(paragraphs)


def build_markdown(report: dict[str, Any], description: str) -> str:
    records = report["record_scale"]
    dedup = report["deduplication_scale"]
    all_text = report["text_scale_all_records"]
    substantive_text = report["text_scale_substantive_records"]

    lines = [
        "# BSRF-HEPS Corpus Scale",
        "",
        description,
        "",
        "## Summary Statistics",
        "",
        "| Metric | Value |",
        "|---|---:|",
        f"| Complete K-V records | {records['complete_kv_records']:,} |",
        f"| Substantive K-V records | {records['substantive_kv_records']:,} |",
        f"| Placeholder records | {records['placeholder_records']:,} |",
        f"| Unique keys | {dedup['keys']['unique_count']:,} |",
        f"| Unique values | {dedup['values']['unique_count']:,} |",
        f"| Unique K-V pairs | {dedup['kv_pairs']['unique_count']:,} |",
        f"| Exact duplicate K-V rows | {dedup['kv_pairs']['extra_duplicate_rows']:,} |",
        f"| Keys with multiple values | {dedup['one_to_many_key_count']:,} |",
        f"| Key-column characters | {all_text['key']['characters']['total']:,} |",
        f"| Value-column characters | {all_text['value']['characters']['total']:,} |",
        f"| Total K-V characters | {all_text['key_value']['characters']['total']:,} |",
        (
            "| K-V characters excluding whitespace | "
            f"{all_text['key_value']['characters']['without_whitespace']:,} |"
        ),
        (
            "| Substantive K-V characters | "
            f"{substantive_text['key_value']['characters']['total']:,} |"
        ),
        f"| Chinese characters | {all_text['key_value']['characters']['chinese']:,} |",
        (
            "| English word units | "
            f"{all_text['key_value']['characters']['english_word_units']:,} |"
        ),
        (
            "| Mean K-V length | "
            f"{all_text['key_value']['length']['mean']:,} characters |"
        ),
        (
            "| Median K-V length | "
            f"{all_text['key_value']['length']['median']:,} characters |"
        ),
        (
            "| K-V length P90 / P95 | "
            f"{all_text['key_value']['length']['p90']:,} / "
            f"{all_text['key_value']['length']['p95']:,} characters |"
        ),
        "",
        (
            "Note: total K-V characters are calculated by summing the key and value "
            'columns. Substantive records exclude placeholder values such as "[Picture '
            'placeholder]" and "——".'
        ),
    ]
    return "\n".join(lines) + "\n"


def print_report(report: dict[str, Any], description: str) -> None:
    records = report["record_scale"]
    dedup = report["deduplication_scale"]
    text = report["text_scale_all_records"]

    print("BSRF-HEPS Corpus size")
    print("=" * 40)
    print(f"completeK-V Record:{records['complete_kv_records']:,}")
    print(f"materialityK-V Record:{records['substantive_kv_records']:,}")
    print(f"onlykey: {dedup['keys']['unique_count']:,}")
    print(f"onlyvalue: {dedup['values']['unique_count']:,}")
    print(f"onlyK-V Right:{dedup['kv_pairs']['unique_count']:,}")
    print(f"K Number of column characters:{text['key']['characters']['total']:,}")
    print(f"V Number of column characters:{text['value']['characters']['total']:,}")
    print(f"K-V Total number of characters:{text['key_value']['characters']['total']:,}")
    print("\nCorpus description:\n")
    print(description)


def main() -> None:
    args = parse_args()
    report = calculate_scale(args.input_file.resolve(), args.sheet)
    description = build_description(report)
    print_report(report, description)

    if not args.no_save:
        json_output = args.json_output.resolve()
        description_output = args.description_output.resolve()
        json_output.parent.mkdir(parents=True, exist_ok=True)
        description_output.parent.mkdir(parents=True, exist_ok=True)
        json_output.write_text(
            json.dumps(report, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        description_output.write_text(
            build_markdown(report, description), encoding="utf-8"
        )


if __name__ == "__main__":
    main()
