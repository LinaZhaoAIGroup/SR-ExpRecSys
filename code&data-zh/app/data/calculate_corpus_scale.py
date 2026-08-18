#!/usr/bin/env python3
"""计算 BSRF-HEPS K-V 语料库规模。

默认读取本脚本同目录下的 BSRF-HEPS-KnowledgeBase-k-v.xlsx，并在终端输出
JSON 格式统计结果。需要保存结果时，可使用 --json-output 指定 JSON 文件路径。
"""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from statistics import mean, median
from typing import Any

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = SCRIPT_DIR / "BSRF-HEPS-KnowledgeBase-k-v.xlsx"
DEFAULT_SHEET = "Sheet1"
DEFAULT_PLACEHOLDERS = ("[图片占位]", "——")
WHITESPACE_RE = re.compile(r"\s+")
ENGLISH_WORD_RE = re.compile(r"[A-Za-z]+(?:[-'][A-Za-z]+)*")
CHINESE_CHAR_RE = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff]")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="计算 Excel K-V 语料库的记录数、去重数和文本规模。"
    )
    parser.add_argument(
        "input_file",
        nargs="?",
        type=Path,
        default=DEFAULT_INPUT,
        help=f"输入 Excel 文件，默认为：{DEFAULT_INPUT.name}",
    )
    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET,
        help=f"待统计工作表，默认为：{DEFAULT_SHEET}",
    )
    parser.add_argument(
        "--json-output",
        type=Path,
        help="可选，将统计结果另存为 JSON 文件",
    )
    parser.add_argument(
        "--include-placeholders",
        action="store_true",
        help="将 [图片占位] 和 —— 视为实质性文本记录",
    )
    return parser.parse_args()


def as_text(value: Any) -> str:
    """将 Excel 单元格转为可统计的文本，同时去除首尾空白。"""
    return "" if value is None else str(value).strip()


def length_stats(values: list[str]) -> dict[str, int | float]:
    lengths = [len(value) for value in values]
    if not lengths:
        return {"total": 0, "mean": 0, "median": 0, "min": 0, "max": 0}
    return {
        "total": sum(lengths),
        "mean": round(mean(lengths), 2),
        "median": median(lengths),
        "min": min(lengths),
        "max": max(lengths),
    }


def count_without_whitespace(values: list[str]) -> int:
    return sum(len(WHITESPACE_RE.sub("", value)) for value in values)


def count_chinese_characters(values: list[str]) -> int:
    return sum(len(CHINESE_CHAR_RE.findall(value)) for value in values)


def count_english_words(values: list[str]) -> int:
    return sum(len(ENGLISH_WORD_RE.findall(value)) for value in values)


def duplicate_summary(values: list[Any]) -> dict[str, int]:
    counts = Counter(values)
    duplicate_groups = {value: count for value, count in counts.items() if count > 1}
    return {
        "duplicate_groups": len(duplicate_groups),
        "rows_in_duplicate_groups": sum(duplicate_groups.values()),
        "extra_duplicate_rows": sum(count - 1 for count in duplicate_groups.values()),
    }


def try_count_tokens(values: list[str]) -> dict[str, Any]:
    """若安装 tiktoken，则额外统计 cl100k_base token 数；否则返回未安装状态。

    cl100k_base 不是 all-MiniLM-L6-v2 的原生 tokenizer，因此该指标只作
    可复现的近似规模参考，不应冒充模型真实 token 数。
    """
    try:
        import tiktoken  # type: ignore
    except ImportError:
        return {"available": False, "encoding": None, "total": None}

    encoding_name = "cl100k_base"
    encoder = tiktoken.get_encoding(encoding_name)
    return {
        "available": True,
        "encoding": encoding_name,
        "total": sum(len(encoder.encode(value)) for value in values),
    }


def inspect_embedding_file(file_path: Path) -> dict[str, Any]:
    """读取现有向量 JSON，核对语句数、向量数及向量维度。"""
    if not file_path.is_file():
        return {"available": False, "file": str(file_path)}

    try:
        payload = json.loads(file_path.read_text(encoding="utf-8"))
        sentences = payload.get("sentences", [])
        embeddings = payload.get("embeddings", [])
        dimensions = sorted(
            {len(embedding) for embedding in embeddings if isinstance(embedding, list)}
        )
        return {
            "available": True,
            "file": str(file_path.resolve()),
            "sentence_count": len(sentences),
            "embedding_count": len(embeddings),
            "dimensions": dimensions,
        }
    except (OSError, UnicodeError, json.JSONDecodeError, AttributeError) as exc:
        return {
            "available": False,
            "file": str(file_path.resolve()),
            "error": str(exc),
        }


def load_kv_records(input_file: Path, sheet_name: str) -> tuple[int, list[dict[str, Any]]]:
    if not input_file.is_file():
        raise FileNotFoundError(f"找不到输入文件：{input_file}")

    workbook = load_workbook(input_file, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"工作表 {sheet_name!r} 不存在，可用工作表：{workbook.sheetnames}"
            )
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ValueError(f"工作表 {sheet_name!r} 为空")

        headers = [as_text(value) for value in header]
        try:
            key_index = headers.index("k")
            value_index = headers.index("v")
        except ValueError as exc:
            raise ValueError(
                f"工作表第一行必须包含 k 和 v 列，实际表头为：{headers}"
            ) from exc

        records = []
        for excel_row, row in enumerate(rows, start=2):
            key = as_text(row[key_index] if key_index < len(row) else None)
            value = as_text(row[value_index] if value_index < len(row) else None)
            if key or value:
                records.append(
                    {"excel_row": excel_row, "key": key, "value": value}
                )
        return worksheet.max_row, records
    finally:
        workbook.close()


def build_report(
    input_file: Path,
    sheet_name: str,
    include_placeholders: bool,
) -> dict[str, Any]:
    physical_max_row, records = load_kv_records(input_file, sheet_name)
    placeholders = set() if include_placeholders else set(DEFAULT_PLACEHOLDERS)

    missing_key_rows = [record for record in records if not record["key"]]
    missing_value_rows = [record for record in records if not record["value"]]
    complete_records = [
        record for record in records if record["key"] and record["value"]
    ]
    substantive_records = [
        record
        for record in complete_records
        if record["value"] not in placeholders
    ]

    keys = [record["key"] for record in complete_records]
    values = [record["value"] for record in complete_records]
    pairs = [(record["key"], record["value"]) for record in complete_records]
    substantive_keys = [record["key"] for record in substantive_records]
    substantive_values = [record["value"] for record in substantive_records]
    substantive_pairs = [
        (record["key"], record["value"]) for record in substantive_records
    ]

    values_by_key: defaultdict[str, set[str]] = defaultdict(set)
    rows_by_key: defaultdict[str, list[int]] = defaultdict(list)
    for record in complete_records:
        values_by_key[record["key"]].add(record["value"])
        rows_by_key[record["key"]].append(record["excel_row"])

    conflicting_keys = {
        key: {
            "row_numbers": row_numbers,
            "value_count": len(values_by_key[key]),
        }
        for key, row_numbers in rows_by_key.items()
        if len(values_by_key[key]) > 1
    }

    combined_text = [key + value for key, value in pairs]
    substantive_combined_text = [key + value for key, value in substantive_pairs]
    placeholder_counts = Counter(
        record["value"]
        for record in complete_records
        if record["value"] in DEFAULT_PLACEHOLDERS
    )

    report: dict[str, Any] = {
        "input_file": str(input_file.resolve()),
        "sheet": sheet_name,
        "physical_max_row": physical_max_row,
        "records": {
            "raw_nonempty_kv_rows": len(records),
            "complete_kv_rows": len(complete_records),
            "substantive_kv_rows": len(substantive_records),
            "missing_key_rows": len(missing_key_rows),
            "missing_value_rows": len(missing_value_rows),
        },
        "uniqueness": {
            "unique_keys": len(set(keys)),
            "unique_values": len(set(values)),
            "unique_kv_pairs": len(set(pairs)),
            "substantive_unique_keys": len(set(substantive_keys)),
            "substantive_unique_values": len(set(substantive_values)),
            "substantive_unique_kv_pairs": len(set(substantive_pairs)),
            "duplicate_keys": duplicate_summary(keys),
            "duplicate_values": duplicate_summary(values),
            "duplicate_kv_pairs": duplicate_summary(pairs),
            "conflicting_key_groups": len(conflicting_keys),
            "conflicting_keys": conflicting_keys,
        },
        "placeholders": {
            "excluded_from_substantive_count": not include_placeholders,
            "markers": list(DEFAULT_PLACEHOLDERS),
            "counts": dict(placeholder_counts),
            "total": sum(placeholder_counts.values()),
        },
        "text_size": {
            "key_characters": length_stats(keys),
            "value_characters": length_stats(values),
            "key_value_characters": length_stats(combined_text),
            "key_characters_without_whitespace": count_without_whitespace(keys),
            "value_characters_without_whitespace": count_without_whitespace(values),
            "key_value_characters_without_whitespace": count_without_whitespace(
                combined_text
            ),
            "key_chinese_characters": count_chinese_characters(keys),
            "value_chinese_characters": count_chinese_characters(values),
            "key_value_chinese_characters": count_chinese_characters(combined_text),
            "key_english_word_units": count_english_words(keys),
            "value_english_word_units": count_english_words(values),
            "substantive_key_value_characters": length_stats(
                substantive_combined_text
            ),
            "substantive_key_value_characters_without_whitespace": count_without_whitespace(
                substantive_combined_text
            ),
            "approximate_tokens": {
                "all_records": try_count_tokens(combined_text),
                "substantive_records": try_count_tokens(substantive_combined_text),
                "note": (
                    "Token 数仅在安装 tiktoken 时计算，并使用 cl100k_base；"
                    "该数值不是 all-MiniLM-L6-v2 的原生 token 数。"
                ),
            },
        },
        "vector_index": {
            "expected_records_from_excel": len(records),
            "model": "all-MiniLM-L6-v2",
            "key_embeddings": inspect_embedding_file(
                input_file.parent / "k_embeddings.json"
            ),
            "value_embeddings": inspect_embedding_file(
                input_file.parent / "v_embeddings.json"
            ),
        },
    }
    return report


def main() -> None:
    args = parse_args()
    report = build_report(args.input_file.resolve(), args.sheet, args.include_placeholders)
    output = json.dumps(report, ensure_ascii=False, indent=2)
    print(output)

    if args.json_output:
        json_path = args.json_output.resolve()
        json_path.parent.mkdir(parents=True, exist_ok=True)
        json_path.write_text(output + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
