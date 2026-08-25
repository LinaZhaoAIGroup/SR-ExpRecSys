#!/usr/bin/env python3
"""完整统计 BSRF-HEPS K-V 语料库的规模。

统计范围仅限语料库规模，包括：记录规模、去重规模、有效文本规模、
字符组成、长度分布以及可选的近似 token 数，不涉及检索或模型性能。
"""

from __future__ import annotations

import argparse
import json
import math
import re
from collections import Counter, defaultdict
from pathlib import Path
from statistics import mean, median
from typing import Any, Iterable

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = SCRIPT_DIR / "BSRF-HEPS-KnowledgeBase-k-v.xlsx"
DEFAULT_SHEET = "Sheet1"
DEFAULT_JSON_OUTPUT = SCRIPT_DIR / "corpus_scale_report.json"
DEFAULT_DESCRIPTION_OUTPUT = SCRIPT_DIR / "语料库规模描述.md"
PLACEHOLDER_VALUES = {"[图片占位]", "——"}

WHITESPACE_RE = re.compile(r"\s")
CHINESE_CHAR_RE = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff]")
ENGLISH_WORD_RE = re.compile(r"[A-Za-z]+(?:[-'][A-Za-z]+)*")
NUMBER_RE = re.compile(r"\d+(?:\.\d+)?")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="统计 Excel K-V 语料库的完整规模并生成论文描述。"
    )
    parser.add_argument(
        "input_file",
        nargs="?",
        type=Path,
        default=DEFAULT_INPUT,
        help=f"输入 Excel 文件，默认为：{DEFAULT_INPUT.name}",
    )
    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET,
        help=f"待统计工作表，默认为：{DEFAULT_SHEET}",
    )
    parser.add_argument(
        "--json-output",
        type=Path,
        default=DEFAULT_JSON_OUTPUT,
        help=f"JSON 报告路径，默认为：{DEFAULT_JSON_OUTPUT.name}",
    )
    parser.add_argument(
        "--description-output",
        type=Path,
        default=DEFAULT_DESCRIPTION_OUTPUT,
        help=f"论文描述路径，默认为：{DEFAULT_DESCRIPTION_OUTPUT.name}",
    )
    parser.add_argument(
        "--no-save",
        action="store_true",
        help="仅在终端输出，不保存 JSON 和 Markdown 文件",
    )
    return parser.parse_args()


def as_text(value: Any) -> str:
    """将 Excel 单元格转换为文本并去除首尾空白。"""
    return "" if value is None else str(value).strip()


def safe_ratio(numerator: int, denominator: int) -> float:
    return round(numerator / denominator, 6) if denominator else 0.0


def percentile(numbers: list[int], proportion: float) -> float:
    """使用线性插值计算分位数。"""
    if not numbers:
        return 0.0
    ordered = sorted(numbers)
    position = (len(ordered) - 1) * proportion
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return float(ordered[lower])
    weight = position - lower
    return ordered[lower] * (1 - weight) + ordered[upper] * weight


def rounded(value: float) -> int | float:
    rounded_value = round(value, 2)
    return int(rounded_value) if rounded_value.is_integer() else rounded_value


def summarize_lengths(texts: list[str]) -> dict[str, int | float]:
    lengths = [len(text) for text in texts]
    if not lengths:
        return {
            "total": 0,
            "mean": 0,
            "median": 0,
            "min": 0,
            "p25": 0,
            "p75": 0,
            "p90": 0,
            "p95": 0,
            "max": 0,
        }
    return {
        "total": sum(lengths),
        "mean": rounded(mean(lengths)),
        "median": rounded(float(median(lengths))),
        "min": min(lengths),
        "p25": rounded(percentile(lengths, 0.25)),
        "p75": rounded(percentile(lengths, 0.75)),
        "p90": rounded(percentile(lengths, 0.90)),
        "p95": rounded(percentile(lengths, 0.95)),
        "max": max(lengths),
    }


def summarize_text(texts: list[str]) -> dict[str, Any]:
    total_characters = sum(len(text) for text in texts)
    whitespace_characters = sum(len(WHITESPACE_RE.findall(text)) for text in texts)
    chinese_characters = sum(len(CHINESE_CHAR_RE.findall(text)) for text in texts)
    english_words = sum(len(ENGLISH_WORD_RE.findall(text)) for text in texts)
    number_sequences = sum(len(NUMBER_RE.findall(text)) for text in texts)

    return {
        "records": len(texts),
        "length": summarize_lengths(texts),
        "characters": {
            "total": total_characters,
            "without_whitespace": total_characters - whitespace_characters,
            "whitespace": whitespace_characters,
            "chinese": chinese_characters,
            "english_word_units": english_words,
            "number_sequences": number_sequences,
        },
    }


def summarize_duplicates(items: Iterable[Any]) -> dict[str, int | float]:
    values = list(items)
    counts = Counter(values)
    repeated_counts = [count for count in counts.values() if count > 1]
    extra_duplicates = sum(count - 1 for count in repeated_counts)
    return {
        "unique_count": len(counts),
        "repeated_groups": len(repeated_counts),
        "rows_in_repeated_groups": sum(repeated_counts),
        "extra_duplicate_rows": extra_duplicates,
        "duplicate_rate": safe_ratio(extra_duplicates, len(values)),
    }


def try_token_count(texts: list[str]) -> dict[str, Any]:
    """可选地用 cl100k_base 估算 token 数，不额外要求安装依赖。"""
    try:
        import tiktoken  # type: ignore
    except ImportError:
        return {
            "available": False,
            "encoding": None,
            "total": None,
            "mean_per_record": None,
            "note": "未安装 tiktoken，未计算近似 token 数。",
        }

    encoding_name = "cl100k_base"
    encoder = tiktoken.get_encoding(encoding_name)
    token_lengths = [len(encoder.encode(text)) for text in texts]
    total = sum(token_lengths)
    return {
        "available": True,
        "encoding": encoding_name,
        "total": total,
        "mean_per_record": rounded(mean(token_lengths)) if token_lengths else 0,
        "note": (
            "该 token 数仅为 cl100k_base 口径的可复现估算，"
            "不代表其他模型 tokenizer 的结果。"
        ),
    }


def load_records(
    input_file: Path, sheet_name: str
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    if not input_file.is_file():
        raise FileNotFoundError(f"找不到输入文件：{input_file}")

    workbook = load_workbook(input_file, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"工作表 {sheet_name!r} 不存在，可用工作表：{workbook.sheetnames}"
            )

        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ValueError(f"工作表 {sheet_name!r} 为空")

        headers = [as_text(value).lower() for value in header]
        try:
            key_index = headers.index("k")
            value_index = headers.index("v")
        except ValueError as exc:
            raise ValueError(
                f"工作表第一行必须包含 k 和 v 列，实际表头为：{headers}"
            ) from exc

        records: list[dict[str, Any]] = []
        blank_rows = 0
        for excel_row, row in enumerate(rows, start=2):
            key = as_text(row[key_index] if key_index < len(row) else None)
            value = as_text(row[value_index] if value_index < len(row) else None)
            if not key and not value:
                blank_rows += 1
                continue
            records.append(
                {
                    "excel_row": excel_row,
                    "key": key,
                    "value": value,
                    "complete": bool(key and value),
                    "placeholder": value in PLACEHOLDER_VALUES,
                }
            )

        metadata = {
            "file": str(input_file.resolve()),
            "file_size_bytes": input_file.stat().st_size,
            "file_size_kib": round(input_file.stat().st_size / 1024, 2),
            "sheet": sheet_name,
            "worksheet_max_row": worksheet.max_row,
            "worksheet_max_column": worksheet.max_column,
            "blank_rows_within_used_range": blank_rows,
        }
        return metadata, records
    finally:
        workbook.close()


def calculate_scale(input_file: Path, sheet_name: str) -> dict[str, Any]:
    metadata, records = load_records(input_file, sheet_name)
    complete_records = [record for record in records if record["complete"]]
    substantive_records = [
        record for record in complete_records if not record["placeholder"]
    ]

    if not complete_records:
        raise ValueError("未找到 k 和 v 均非空的完整语料记录")

    keys = [record["key"] for record in complete_records]
    values = [record["value"] for record in complete_records]
    pairs = [(record["key"], record["value"]) for record in complete_records]
    # K-V 总字符量采用 K 字符数与 V 字符数直接相加的统计口径。
    combined = [key + value for key, value in pairs]

    substantive_keys = [record["key"] for record in substantive_records]
    substantive_values = [record["value"] for record in substantive_records]
    substantive_pairs = [
        (record["key"], record["value"]) for record in substantive_records
    ]
    substantive_combined = [key + value for key, value in substantive_pairs]

    values_by_key: defaultdict[str, set[str]] = defaultdict(set)
    for key, value in pairs:
        values_by_key[key].add(value)
    one_to_many_keys = {
        key: len(distinct_values)
        for key, distinct_values in values_by_key.items()
        if len(distinct_values) > 1
    }

    placeholder_counts = Counter(
        record["value"] for record in complete_records if record["placeholder"]
    )
    missing_key_rows = [record["excel_row"] for record in records if not record["key"]]
    missing_value_rows = [
        record["excel_row"] for record in records if not record["value"]
    ]

    return {
        "metadata": metadata,
        "record_scale": {
            "nonempty_data_rows": len(records),
            "complete_kv_records": len(complete_records),
            "incomplete_records": len(records) - len(complete_records),
            "missing_key_rows": missing_key_rows,
            "missing_value_rows": missing_value_rows,
            "substantive_kv_records": len(substantive_records),
            "substantive_record_rate": safe_ratio(
                len(substantive_records), len(complete_records)
            ),
            "placeholder_records": len(complete_records) - len(substantive_records),
            "placeholder_rate": safe_ratio(
                len(complete_records) - len(substantive_records),
                len(complete_records),
            ),
            "placeholder_counts": dict(placeholder_counts),
        },
        "deduplication_scale": {
            "keys": summarize_duplicates(keys),
            "values": summarize_duplicates(values),
            "kv_pairs": summarize_duplicates(pairs),
            "one_to_many_key_count": len(one_to_many_keys),
            "one_to_many_key_rate": safe_ratio(len(one_to_many_keys), len(set(keys))),
            "substantive_unique_keys": len(set(substantive_keys)),
            "substantive_unique_values": len(set(substantive_values)),
            "substantive_unique_kv_pairs": len(set(substantive_pairs)),
        },
        "text_scale_all_records": {
            "key": summarize_text(keys),
            "value": summarize_text(values),
            "key_value": summarize_text(combined),
            "approximate_tokens": try_token_count(combined),
        },
        "text_scale_substantive_records": {
            "key": summarize_text(substantive_keys),
            "value": summarize_text(substantive_values),
            "key_value": summarize_text(substantive_combined),
            "approximate_tokens": try_token_count(substantive_combined),
        },
    }


def build_description(report: dict[str, Any]) -> str:
    records = report["record_scale"]
    dedup = report["deduplication_scale"]
    all_text = report["text_scale_all_records"]
    substantive_text = report["text_scale_substantive_records"]

    kv = all_text["key_value"]
    key = all_text["key"]
    value = all_text["value"]
    substantive_kv = substantive_text["key_value"]

    paragraphs = [
        (
            "本研究构建的 BSRF-HEPS 键值型语料库共包含 "
            f"{records['complete_kv_records']} 条完整 K-V 语料记录。"
            f"其中包含 {dedup['keys']['unique_count']} 个唯一 key、"
            f"{dedup['values']['unique_count']} 个唯一 value 和 "
            f"{dedup['kv_pairs']['unique_count']} 个唯一 K-V 对；"
            f"完全重复 K-V 记录为 {dedup['kv_pairs']['extra_duplicate_rows']} 条，"
            f"重复率为 {dedup['kv_pairs']['duplicate_rate']:.2%}。"
            f"此外，共有 {dedup['one_to_many_key_count']} 个 key 对应多个不同 value。"
        ),
        (
            f"在文本规模方面，K 列共包含 {key['characters']['total']:,} 个字符，"
            f"V 列共包含 {value['characters']['total']:,} 个字符；"
            f"将 K 与 V 的字符量直接合计，语料总规模为 "
            f"{kv['characters']['total']:,} 个字符，去除空白字符后为 "
            f"{kv['characters']['without_whitespace']:,} 个字符。"
            f"每条 K-V 记录平均包含 {kv['length']['mean']} 个字符，"
            f"中位数为 {kv['length']['median']} 个字符，"
            f"长度范围为 {kv['length']['min']}-{kv['length']['max']} 个字符，"
            f"P90 和 P95 分别为 {kv['length']['p90']} 和 {kv['length']['p95']} 个字符。"
        ),
        (
            f"语料中共识别出 {kv['characters']['chinese']:,} 个中文字符、"
            f"{kv['characters']['english_word_units']:,} 个英文词单元和 "
            f"{kv['characters']['number_sequences']:,} 个数字序列。"
            f"排除“[图片占位]”和“——”后，共保留 "
            f"{records['substantive_kv_records']} 条实质性文本记录，"
            f"占完整记录的 {records['substantive_record_rate']:.2%}；"
            f"其实质性 K-V 文本规模为 "
            f"{substantive_kv['characters']['total']:,} 个字符。"
        ),
    ]
    return "\n\n".join(paragraphs)


def build_markdown(report: dict[str, Any], description: str) -> str:
    records = report["record_scale"]
    dedup = report["deduplication_scale"]
    all_text = report["text_scale_all_records"]
    substantive_text = report["text_scale_substantive_records"]

    lines = [
        "# BSRF-HEPS 语料库规模描述",
        "",
        description,
        "",
        "## 统计口径",
        "",
        "| 指标 | 数值 |",
        "|---|---:|",
        f"| 完整 K-V 记录 | {records['complete_kv_records']:,} |",
        f"| 实质性 K-V 记录 | {records['substantive_kv_records']:,} |",
        f"| 占位记录 | {records['placeholder_records']:,} |",
        f"| 唯一 key | {dedup['keys']['unique_count']:,} |",
        f"| 唯一 value | {dedup['values']['unique_count']:,} |",
        f"| 唯一 K-V 对 | {dedup['kv_pairs']['unique_count']:,} |",
        f"| 完全重复 K-V 记录 | {dedup['kv_pairs']['extra_duplicate_rows']:,} |",
        f"| 一对多 key | {dedup['one_to_many_key_count']:,} |",
        f"| K 列字符数 | {all_text['key']['characters']['total']:,} |",
        f"| V 列字符数 | {all_text['value']['characters']['total']:,} |",
        f"| K-V 总字符数 | {all_text['key_value']['characters']['total']:,} |",
        (
            "| K-V 去空白字符数 | "
            f"{all_text['key_value']['characters']['without_whitespace']:,} |"
        ),
        (
            "| 实质性 K-V 字符数 | "
            f"{substantive_text['key_value']['characters']['total']:,} |"
        ),
        f"| 中文字符数 | {all_text['key_value']['characters']['chinese']:,} |",
        (
            "| 英文词单元数 | "
            f"{all_text['key_value']['characters']['english_word_units']:,} |"
        ),
        (
            "| 平均 K-V 长度 | "
            f"{all_text['key_value']['length']['mean']:,} 字符 |"
        ),
        (
            "| K-V 长度中位数 | "
            f"{all_text['key_value']['length']['median']:,} 字符 |"
        ),
        (
            "| K-V 长度 P90 / P95 | "
            f"{all_text['key_value']['length']['p90']:,} / "
            f"{all_text['key_value']['length']['p95']:,} 字符 |"
        ),
        "",
        (
            "注：K-V 总字符数按 K 列与 V 列字符量直接相加计算；"
            "实质性记录排除了 value 为“[图片占位]”或“——”的条目。"
        ),
    ]
    return "\n".join(lines) + "\n"


def print_report(report: dict[str, Any], description: str) -> None:
    records = report["record_scale"]
    dedup = report["deduplication_scale"]
    text = report["text_scale_all_records"]

    print("BSRF-HEPS 语料库规模")
    print("=" * 40)
    print(f"完整 K-V 记录：{records['complete_kv_records']:,}")
    print(f"实质性 K-V 记录：{records['substantive_kv_records']:,}")
    print(f"唯一 key：{dedup['keys']['unique_count']:,}")
    print(f"唯一 value：{dedup['values']['unique_count']:,}")
    print(f"唯一 K-V 对：{dedup['kv_pairs']['unique_count']:,}")
    print(f"K 列字符数：{text['key']['characters']['total']:,}")
    print(f"V 列字符数：{text['value']['characters']['total']:,}")
    print(f"K-V 总字符数：{text['key_value']['characters']['total']:,}")
    print("\n语料库描述：\n")
    print(description)


def main() -> None:
    args = parse_args()
    report = calculate_scale(args.input_file.resolve(), args.sheet)
    description = build_description(report)
    print_report(report, description)

    if not args.no_save:
        json_output = args.json_output.resolve()
        description_output = args.description_output.resolve()
        json_output.parent.mkdir(parents=True, exist_ok=True)
        description_output.parent.mkdir(parents=True, exist_ok=True)
        json_output.write_text(
            json.dumps(report, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        description_output.write_text(
            build_markdown(report, description), encoding="utf-8"
        )


if __name__ == "__main__":
    main()
