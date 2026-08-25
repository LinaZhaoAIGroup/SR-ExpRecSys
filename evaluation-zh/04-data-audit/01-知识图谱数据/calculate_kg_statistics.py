#!/usr/bin/env python3
"""Audit the size and structural consistency of the BSRF-HEPS KG.

The script is read-only with respect to the source CSV/XLSX files. It reports
raw rows, complete triples, exact duplicate triples, unique nodes/edges,
entity-type and relation-label distributions, and several quality flags. It
also includes the size of the K-V source corpus when that file is available.

Important: these structural statistics do not measure extraction precision or
recall. Use prepare_kg_validation_sample.py and
calculate_kg_validation_metrics.py for expert validation.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_MAIN_KG = SCRIPT_DIR / "02--全部KG" / "BSRF_HEPS_tuple.xlsx"
DEFAULT_MAIN_SHEET = "Sheet1"
DEFAULT_EXPERIMENTAL_KG = (
    SCRIPT_DIR / "01--实验技术KG" / "experimental_technique_tuple.csv"
)
DEFAULT_CORPUS = (
    SCRIPT_DIR.parent / "00-语料库规模" / "BSRF-HEPS-KnowledgeBase-k-v.xlsx"
)
DEFAULT_CORPUS_SHEET = "Sheet1"
DEFAULT_OUTPUT_DIR = SCRIPT_DIR / "outputs" / "kg_statistics"

TRIPLE_COLUMNS = (
    "sub_name",
    "sub_type",
    "rel_type",
    "rel_name",
    "obj_name",
    "obj_type",
)
TRIPLE_KEY_COLUMNS = ("sub_name", "rel_name", "obj_name")
PLACEHOLDER_VALUES = {"[图片占位]", "——"}
WHITESPACE_RE = re.compile(r"\s+")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="统计知识图谱规模，并输出关系/实体类别分布和质量审计明细。"
    )
    parser.add_argument("--main-kg", type=Path, default=DEFAULT_MAIN_KG)
    parser.add_argument("--main-sheet", default=DEFAULT_MAIN_SHEET)
    parser.add_argument(
        "--experimental-kg", type=Path, default=DEFAULT_EXPERIMENTAL_KG
    )
    parser.add_argument("--skip-experimental", action="store_true")
    parser.add_argument("--corpus", type=Path, default=DEFAULT_CORPUS)
    parser.add_argument("--corpus-sheet", default=DEFAULT_CORPUS_SHEET)
    parser.add_argument("--skip-corpus", action="store_true")
    parser.add_argument("--top-relations", type=int, default=15)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args()


def normalize_text(value: Any) -> str:
    """Normalize cell text without applying semantic synonym merging."""
    if value is None:
        return ""
    return WHITESPACE_RE.sub(" ", str(value)).strip()


def normalized_header(value: Any) -> str:
    return normalize_text(value).lstrip("\ufeff").lower()


def load_xlsx_table(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    if not path.is_file():
        raise FileNotFoundError(f"找不到 Excel 文件：{path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"工作表 {sheet_name!r} 不存在；可用工作表：{workbook.sheetnames}"
            )
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ValueError(f"工作表 {sheet_name!r} 为空")
        all_headers = [normalized_header(value) for value in header]
        header_positions = [
            (index, header_name)
            for index, header_name in enumerate(all_headers)
            if header_name
        ]
        if not header_positions:
            raise ValueError(f"工作表 {sheet_name!r} 没有有效表头")

        headers = [header_name for _, header_name in header_positions]
        duplicate_headers = sorted(
            {header_name for header_name in headers if headers.count(header_name) > 1}
        )
        if duplicate_headers:
            raise ValueError(f"表头存在重复列：{duplicate_headers}")

        records: list[dict[str, Any]] = []
        for excel_row, row in enumerate(rows, start=2):
            record = {
                header_name: normalize_text(row[index] if index < len(row) else None)
                for index, header_name in header_positions
            }
            if not any(record.values()):
                continue
            record["source_row"] = excel_row
            records.append(record)
        return records
    finally:
        workbook.close()


def load_csv_table(path: Path) -> list[dict[str, Any]]:
    if not path.is_file():
        raise FileNotFoundError(f"找不到 CSV 文件：{path}")

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise ValueError(f"CSV 文件没有表头：{path}")
        header_map = {name: normalized_header(name) for name in reader.fieldnames}
        records: list[dict[str, Any]] = []
        for source_row, row in enumerate(reader, start=2):
            record = {
                header_map[name]: normalize_text(row.get(name))
                for name in reader.fieldnames
            }
            if not any(record.values()):
                continue
            record["source_row"] = source_row
            records.append(record)
        return records


def load_table(path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return load_csv_table(path)
    if suffix in {".xlsx", ".xlsm"}:
        if sheet_name is None:
            raise ValueError("读取 Excel 时必须指定工作表名称")
        return load_xlsx_table(path, sheet_name)
    raise ValueError(f"不支持的文件类型：{path.suffix}")


def validate_triple_columns(records: list[dict[str, Any]], source: Path) -> None:
    if not records:
        raise ValueError(f"未读取到任何数据：{source}")
    missing = [column for column in TRIPLE_COLUMNS if column not in records[0]]
    if missing:
        raise ValueError(f"{source.name} 缺少三元组字段：{missing}")


def triple_key(record: dict[str, Any]) -> tuple[str, str, str]:
    return tuple(record[column] for column in TRIPLE_KEY_COLUMNS)  # type: ignore[return-value]


def sorted_count_rows(counter: Counter[str], label_column: str) -> list[dict[str, Any]]:
    return [
        {label_column: label, "count": count}
        for label, count in sorted(counter.items(), key=lambda item: (-item[1], item[0]))
    ]


def audit_triples(
    records: list[dict[str, Any]], dataset_name: str, source: Path, sheet: str | None
) -> dict[str, Any]:
    validate_triple_columns(records, source)

    incomplete = [
        record
        for record in records
        if any(not record[column] for column in TRIPLE_KEY_COLUMNS)
    ]
    complete = [
        record
        for record in records
        if all(record[column] for column in TRIPLE_KEY_COLUMNS)
    ]

    rows_by_key: defaultdict[tuple[str, str, str], list[int]] = defaultdict(list)
    first_record_by_key: dict[tuple[str, str, str], dict[str, Any]] = {}
    for record in complete:
        key = triple_key(record)
        rows_by_key[key].append(record["source_row"])
        first_record_by_key.setdefault(key, record)
    unique_records = list(first_record_by_key.values())

    duplicate_groups = []
    for key, source_rows in rows_by_key.items():
        if len(source_rows) > 1:
            duplicate_groups.append(
                {
                    "sub_name": key[0],
                    "rel_name": key[1],
                    "obj_name": key[2],
                    "occurrences": len(source_rows),
                    "extra_duplicate_rows": len(source_rows) - 1,
                    "source_rows": ";".join(str(value) for value in source_rows),
                }
            )

    nodes: set[str] = set()
    typed_nodes: set[tuple[str, str]] = set()
    class_to_nodes: defaultdict[str, set[str]] = defaultdict(set)
    types_by_node: defaultdict[str, set[str]] = defaultdict(set)
    relation_name_counts: Counter[str] = Counter()
    relation_type_counts: Counter[str] = Counter()
    relation_pair_counts: Counter[tuple[str, str]] = Counter()
    self_loops: list[dict[str, Any]] = []
    relation_mismatches: list[dict[str, Any]] = []
    placeholder_rows: list[dict[str, Any]] = []

    for record in unique_records:
        for name_column, type_column in (
            ("sub_name", "sub_type"),
            ("obj_name", "obj_type"),
        ):
            name = record[name_column]
            entity_type = record[type_column]
            if name:
                nodes.add(name)
            if name and entity_type:
                typed_nodes.add((name, entity_type))
                class_to_nodes[entity_type].add(name)
                types_by_node[name].add(entity_type)

        relation_name_counts[record["rel_name"]] += 1
        relation_type_counts[record["rel_type"]] += 1
        relation_pair_counts[(record["rel_type"], record["rel_name"])] += 1

        if record["sub_name"] == record["obj_name"]:
            self_loops.append(record)
        if record["rel_type"] != record["rel_name"]:
            relation_mismatches.append(record)
        if record["sub_name"] in PLACEHOLDER_VALUES or record["obj_name"] in PLACEHOLDER_VALUES:
            placeholder_rows.append(record)

    type_conflicts = [
        {
            "entity_name": name,
            "type_count": len(entity_types),
            "entity_types": ";".join(sorted(entity_types)),
        }
        for name, entity_types in sorted(types_by_node.items())
        if len(entity_types) > 1
    ]

    entity_class_rows = [
        {"entity_type": entity_type, "unique_node_count": len(class_to_nodes[entity_type])}
        for entity_type in sorted(
            class_to_nodes, key=lambda value: (-len(class_to_nodes[value]), value)
        )
    ]
    relation_name_rows = sorted_count_rows(relation_name_counts, "relation_name")
    relation_type_rows = sorted_count_rows(relation_type_counts, "relation_type")
    relation_mapping_rows = [
        {
            "raw_rel_type": relation_type,
            "raw_rel_name": relation_name,
            "unique_triple_count": count,
            "canonical_relation": "",
            "decision_keep_merge_remove": "",
            "notes": "",
        }
        for (relation_type, relation_name), count in sorted(
            relation_pair_counts.items(), key=lambda item: (-item[1], item[0])
        )
    ]
    entity_mapping_rows = [
        {
            "raw_entity_type": row["entity_type"],
            "unique_node_count": row["unique_node_count"],
            "canonical_entity_class": "",
            "decision_keep_merge_remove": "",
            "notes": "",
        }
        for row in entity_class_rows
    ]

    quality_flags: list[dict[str, Any]] = []
    for record in incomplete:
        quality_flags.append(
            quality_row("incomplete_triple", record, "subject/relation/object 至少一项为空")
        )
    for group in duplicate_groups:
        quality_flags.append(
            {
                "issue": "exact_duplicate_group",
                "source_row": group["source_rows"],
                "sub_name": group["sub_name"],
                "rel_type": "",
                "rel_name": group["rel_name"],
                "obj_name": group["obj_name"],
                "detail": f"出现 {group['occurrences']} 次",
            }
        )
    for record in self_loops:
        quality_flags.append(quality_row("self_loop_candidate", record, "subject 与 object 相同"))
    for record in relation_mismatches:
        quality_flags.append(
            quality_row("relation_label_mismatch", record, "rel_type 与 rel_name 不相同")
        )
    for record in placeholder_rows:
        quality_flags.append(
            quality_row("placeholder_entity", record, "subject 或 object 为占位文本")
        )
    for conflict in type_conflicts:
        quality_flags.append(
            {
                "issue": "entity_type_conflict",
                "source_row": "",
                "sub_name": conflict["entity_name"],
                "rel_type": "",
                "rel_name": "",
                "obj_name": "",
                "detail": conflict["entity_types"],
            }
        )

    summary = {
        "dataset": dataset_name,
        "source_file": str(source.resolve()),
        "sheet": sheet,
        "raw_nonblank_rows": len(records),
        "complete_triple_rows": len(complete),
        "incomplete_triple_rows": len(incomplete),
        "unique_directed_triples_edges": len(unique_records),
        "exact_duplicate_groups": len(duplicate_groups),
        "extra_exact_duplicate_rows": sum(
            row["extra_duplicate_rows"] for row in duplicate_groups
        ),
        "unique_node_names": len(nodes),
        "unique_typed_node_assignments": len(typed_nodes),
        "observed_entity_type_labels": len(class_to_nodes),
        "entity_names_with_multiple_types": len(type_conflicts),
        "observed_relation_name_labels": len(relation_name_counts),
        "observed_relation_type_labels": len(relation_type_counts),
        "relation_type_name_mismatch_edges": len(relation_mismatches),
        "self_loop_candidates": len(self_loops),
        "placeholder_entity_edges": len(placeholder_rows),
    }

    return {
        "summary": summary,
        "entity_class_distribution": entity_class_rows,
        "relation_name_distribution": relation_name_rows,
        "relation_type_distribution": relation_type_rows,
        "duplicate_groups": duplicate_groups,
        "entity_type_conflicts": type_conflicts,
        "quality_flags": quality_flags,
        "relation_mapping_template": relation_mapping_rows,
        "entity_type_mapping_template": entity_mapping_rows,
    }


def quality_row(issue: str, record: dict[str, Any], detail: str) -> dict[str, Any]:
    return {
        "issue": issue,
        "source_row": record.get("source_row", ""),
        "sub_name": record.get("sub_name", ""),
        "rel_type": record.get("rel_type", ""),
        "rel_name": record.get("rel_name", ""),
        "obj_name": record.get("obj_name", ""),
        "detail": detail,
    }


def audit_corpus(path: Path, sheet_name: str) -> dict[str, Any]:
    records = load_xlsx_table(path, sheet_name)
    if not records:
        raise ValueError(f"语料库为空：{path}")
    missing = [column for column in ("k", "v") if column not in records[0]]
    if missing:
        raise ValueError(f"语料库缺少字段：{missing}")

    complete = [record for record in records if record["k"] and record["v"]]
    substantive = [
        record for record in complete if record["v"] not in PLACEHOLDER_VALUES
    ]
    unique_pairs = {(record["k"], record["v"]) for record in complete}
    return {
        "source_file": str(path.resolve()),
        "sheet": sheet_name,
        "confirmed_as_complete_kg_extraction_source": False,
        "confirmation_note": (
            "The KG table contains no source provenance. Confirm that this K-V file is the "
            "complete extraction source before reporting it as the source-corpus size."
        ),
        "complete_kv_records": len(complete),
        "unique_kv_pairs": len(unique_pairs),
        "substantive_kv_records": len(substantive),
        "placeholder_records": len(complete) - len(substantive),
        "total_kv_characters": sum(
            len(record["k"]) + len(record["v"]) for record in complete
        ),
        "substantive_kv_characters": sum(
            len(record["k"]) + len(record["v"]) for record in substantive
        ),
    }


def write_csv(path: Path, rows: Iterable[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_dataset_outputs(
    output_dir: Path, prefix: str, report: dict[str, Any]
) -> None:
    write_csv(
        output_dir / f"{prefix}_entity_class_distribution.csv",
        report["entity_class_distribution"],
        ["entity_type", "unique_node_count"],
    )
    write_csv(
        output_dir / f"{prefix}_relation_name_distribution.csv",
        report["relation_name_distribution"],
        ["relation_name", "count"],
    )
    write_csv(
        output_dir / f"{prefix}_relation_type_distribution.csv",
        report["relation_type_distribution"],
        ["relation_type", "count"],
    )
    write_csv(
        output_dir / f"{prefix}_quality_flags.csv",
        report["quality_flags"],
        [
            "issue",
            "source_row",
            "sub_name",
            "rel_type",
            "rel_name",
            "obj_name",
            "detail",
        ],
    )
    write_csv(
        output_dir / f"{prefix}_relation_mapping_template.csv",
        report["relation_mapping_template"],
        [
            "raw_rel_type",
            "raw_rel_name",
            "unique_triple_count",
            "canonical_relation",
            "decision_keep_merge_remove",
            "notes",
        ],
    )
    write_csv(
        output_dir / f"{prefix}_entity_type_mapping_template.csv",
        report["entity_type_mapping_template"],
        [
            "raw_entity_type",
            "unique_node_count",
            "canonical_entity_class",
            "decision_keep_merge_remove",
            "notes",
        ],
    )


def markdown_table(headers: list[str], rows: list[list[Any]]) -> list[str]:
    lines = [
        "| " + " | ".join(headers) + " |",
        "|" + "|".join("---" for _ in headers) + "|",
    ]
    for row in rows:
        lines.append("| " + " | ".join(str(value) for value in row) + " |")
    return lines


def build_markdown(
    reports: dict[str, dict[str, Any]],
    corpus: dict[str, Any] | None,
    top_relations: int,
) -> str:
    lines = [
        "# BSRF-HEPS 知识图谱规模与结构审计",
        "",
        "## 统计口径",
        "",
        "- 主知识图谱默认采用 `BSRF_HEPS_tuple.xlsx` 的 `Sheet1`；`Sheet1 (2)` 是网站旧版子集，不与主表相加。",
        "- 唯一边按标准化空白后的 `(sub_name, rel_name, obj_name)` 去重。",
        "- 唯一节点按实体名称计数，与当前图数据库加载程序按 `name` 去重的行为一致。",
        "- 带类型节点数按唯一 `(实体名称, 实体类型)` 计数；同名实体可能对应多个类型，因此该数可大于唯一节点数。",
        "- 关系名称和实体类型均按数据表中的原始标签计数，未擅自进行语义合并。",
        "- 本报告是结构审计，不代表抽取精确率、召回率或专家一致性。",
        "",
        "## 总体统计",
        "",
    ]

    table_rows: list[list[Any]] = []
    for report in reports.values():
        summary = report["summary"]
        table_rows.append(
            [
                summary["dataset"],
                summary["raw_nonblank_rows"],
                summary["complete_triple_rows"],
                summary["unique_directed_triples_edges"],
                summary["unique_node_names"],
                summary["observed_entity_type_labels"],
                summary["observed_relation_name_labels"],
            ]
        )
    lines.extend(
        markdown_table(
            [
                "数据集",
                "原始非空行",
                "完整三元组行",
                "唯一边",
                "唯一节点",
                "原始实体类型标签",
                "原始关系名称标签",
            ],
            table_rows,
        )
    )

    if corpus is not None:
        lines.extend(
            [
                "",
                "## 候选源语料库文件规模（需作者确认）",
                "",
            ]
        )
        lines.extend(
            markdown_table(
                ["指标", "数值"],
                [
                    ["完整 K-V 记录", corpus["complete_kv_records"]],
                    ["唯一 K-V 对", corpus["unique_kv_pairs"]],
                    ["实质性 K-V 记录", corpus["substantive_kv_records"]],
                    ["占位记录", corpus["placeholder_records"]],
                    ["K-V 总字符数", f"{corpus['total_kv_characters']:,}"],
                    [
                        "实质性 K-V 字符数",
                        f"{corpus['substantive_kv_characters']:,}",
                    ],
                ],
            )
        )
        lines.extend(
            [
                "",
                "注：知识图谱表没有来源文档或段落标识，脚本无法验证该 K-V 文件是否为完整抽取来源。作者确认前，不应把这些数字直接表述为知识图谱的 source-corpus size。",
            ]
        )

    for key, report in reports.items():
        summary = report["summary"]
        lines.extend(
            [
                "",
                f"## {summary['dataset']}",
                "",
                f"质量审计：额外精确重复行 {summary['extra_exact_duplicate_rows']} 条；"
                f"同名多类型实体 {summary['entity_names_with_multiple_types']} 个；"
                f"自环候选 {summary['self_loop_candidates']} 条；"
                f"`rel_type` 与 `rel_name` 不一致的唯一边 "
                f"{summary['relation_type_name_mismatch_edges']} 条。",
                "",
                "### 实体类型分布",
                "",
            ]
        )
        lines.extend(
            markdown_table(
                ["原始实体类型", "唯一节点数"],
                [
                    [row["entity_type"] or "[空]", row["unique_node_count"]]
                    for row in report["entity_class_distribution"]
                ],
            )
        )
        lines.extend(["", f"### 前 {top_relations} 个关系名称", ""])
        lines.extend(
            markdown_table(
                ["原始关系名称", "唯一边数"],
                [
                    [row["relation_name"] or "[空]", row["count"]]
                    for row in report["relation_name_distribution"][:top_relations]
                ],
            )
        )
        remaining = report["relation_name_distribution"][top_relations:]
        if remaining:
            lines.append("")
            lines.append(
                f"其余 {len(remaining)} 个关系名称合计 "
                f"{sum(row['count'] for row in remaining)} 条唯一边；完整分布见 CSV 输出。"
            )

    lines.extend(
        [
            "",
            "## 解释限制",
            "",
            "当前主表中原始实体类型和关系标签明显多于论文所述核心本体类别，且存在同名实体多类型、自环和重复记录。应先由领域专家完成实体类型/关系映射与异常 adjudication，再把最终类别数和边数写入论文。仅凭最终三元组表无法恢复同义词合并、人工纠正和删除的历史数量。",
            "",
        ]
    )
    return "\n".join(lines)


def main() -> None:
    args = parse_args()
    output_dir = args.output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    reports: dict[str, dict[str, Any]] = {}
    main_records = load_table(args.main_kg.resolve(), args.main_sheet)
    reports["main"] = audit_triples(
        main_records,
        dataset_name="主知识图谱",
        source=args.main_kg.resolve(),
        sheet=args.main_sheet,
    )

    if not args.skip_experimental and args.experimental_kg.is_file():
        experimental_records = load_table(args.experimental_kg.resolve())
        reports["experimental"] = audit_triples(
            experimental_records,
            dataset_name="实验技术专项图谱",
            source=args.experimental_kg.resolve(),
            sheet=None,
        )

    corpus = None
    if not args.skip_corpus and args.corpus.is_file():
        corpus = audit_corpus(args.corpus.resolve(), args.corpus_sheet)

    for prefix, report in reports.items():
        write_dataset_outputs(output_dir, prefix, report)

    complete_report = {
        "definitions": {
            "edge_identity": list(TRIPLE_KEY_COLUMNS),
            "node_identity": "normalized entity name",
            "typed_node_identity": ["entity name", "entity type"],
            "semantic_synonym_merging_applied": False,
        },
        "source_corpus": corpus,
        "datasets": reports,
    }
    (output_dir / "kg_statistics_report.json").write_text(
        json.dumps(complete_report, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (output_dir / "KG_statistics_summary.md").write_text(
        build_markdown(reports, corpus, args.top_relations), encoding="utf-8"
    )

    print(f"统计完成：{output_dir}")


if __name__ == "__main__":
    main()
