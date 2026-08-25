#!/usr/bin/env python3
"""Create reproducible, blinded CSV forms for two-expert KG validation.

The script samples unique complete triples from the main KG with a fixed random
seed. It does not alter the KG. Two identical annotation files are created so
that experts can work independently before adjudication.

Because the current triple table has no source-document/provenance columns,
reviewers must locate and record the supporting source passage during review.
Without source evidence, extraction correctness cannot be judged reliably.
"""

from __future__ import annotations

import argparse
import csv
import random
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = SCRIPT_DIR / "02--全部KG" / "BSRF_HEPS_tuple.xlsx"
DEFAULT_SHEET = "Sheet1"
DEFAULT_OUTPUT_DIR = SCRIPT_DIR / "validation" / "round_1"
DEFAULT_SAMPLE_SIZE = 235
DEFAULT_SEED = 20260812
TRIPLE_COLUMNS = (
    "sub_name",
    "sub_type",
    "rel_type",
    "rel_name",
    "obj_name",
    "obj_type",
)
TRIPLE_KEY_COLUMNS = ("sub_name", "rel_name", "obj_name")
WHITESPACE_RE = re.compile(r"\s+")

ANNOTATION_COLUMNS = [
    "sample_id",
    "source_file",
    "source_sheet",
    "source_row",
    "sub_name",
    "sub_type",
    "rel_type",
    "rel_name",
    "obj_name",
    "obj_type",
    "facility_BSRF_or_HEPS",
    "source_document",
    "source_location",
    "source_excerpt",
    "source_support_Y_N_UNCERTAIN",
    "subject_correct_Y_N_UNCERTAIN",
    "relation_correct_Y_N_UNCERTAIN",
    "object_correct_Y_N_UNCERTAIN",
    "entity_types_correct_Y_N_UNCERTAIN",
    "triple_correct_Y_N_UNCERTAIN",
    "error_types_semicolon_separated",
    "action_KEEP_CORRECT_REMOVE_UNCERTAIN",
    "corrected_sub_name",
    "corrected_sub_type",
    "corrected_rel_type",
    "corrected_rel_name",
    "corrected_obj_name",
    "corrected_obj_type",
    "notes",
]

ERROR_TYPE_CODES = [
    "UNSUPPORTED_TRIPLE",
    "WRONG_SUBJECT",
    "WRONG_RELATION",
    "WRONG_OBJECT",
    "WRONG_ENTITY_TYPE",
    "OVERLY_BROAD_OR_NARROW",
    "DUPLICATE",
    "SYNONYM_NOT_MERGED",
    "SELF_LOOP",
    "INCONSISTENT_SCHEMA",
    "MISSING_PROVENANCE",
    "OTHER",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="从主知识图谱中随机抽取唯一三元组，生成两位专家的独立盲评表。"
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--sample-size", type=int, default=DEFAULT_SAMPLE_SIZE)
    parser.add_argument("--seed", type=int, default=DEFAULT_SEED)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args()


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return WHITESPACE_RE.sub(" ", str(value)).strip()


def load_unique_triples(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    if not path.is_file():
        raise FileNotFoundError(f"找不到知识图谱文件：{path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"工作表 {sheet_name!r} 不存在；可用工作表：{workbook.sheetnames}"
            )
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ValueError(f"工作表 {sheet_name!r} 为空")
        headers = [normalize_text(value).lstrip("\ufeff").lower() for value in header]
        missing = [column for column in TRIPLE_COLUMNS if column not in headers]
        if missing:
            raise ValueError(f"缺少三元组字段：{missing}")

        unique: dict[tuple[str, str, str], dict[str, Any]] = {}
        for source_row, row in enumerate(rows, start=2):
            values = list(row) + [None] * max(0, len(headers) - len(row))
            record = {
                name: normalize_text(values[index])
                for index, name in enumerate(headers)
            }
            if not any(record.get(column, "") for column in TRIPLE_COLUMNS):
                continue
            if any(not record.get(column, "") for column in TRIPLE_KEY_COLUMNS):
                continue
            key = tuple(record[column] for column in TRIPLE_KEY_COLUMNS)
            if key not in unique:
                record["source_row"] = source_row
                unique[key] = record
        return list(unique.values())
    finally:
        workbook.close()


def make_annotation_rows(
    sampled: list[dict[str, Any]], source_file: Path, source_sheet: str
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for index, record in enumerate(sampled, start=1):
        output = {column: "" for column in ANNOTATION_COLUMNS}
        output.update(
            {
                "sample_id": f"KG-{index:04d}",
                "source_file": str(source_file.resolve()),
                "source_sheet": source_sheet,
                "source_row": record["source_row"],
            }
        )
        for column in TRIPLE_COLUMNS:
            output[column] = record[column]
        rows.append(output)
    return rows


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=ANNOTATION_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def write_protocol(path: Path, population: int, sample_size: int, seed: int) -> None:
    error_codes = "、".join(f"`{code}`" for code in ERROR_TYPE_CODES)
    if population > 1:
        finite_population_correction = (population - sample_size) / (population - 1)
        worst_case_margin = 1.95996398454 * (
            0.25 / sample_size * finite_population_correction
        ) ** 0.5
    else:
        worst_case_margin = 0.0
    path.write_text(
        f"""# 知识图谱双专家验证说明

## 抽样信息

- 抽样总体：主知识图谱中按 `(sub_name, rel_name, obj_name)` 去重后的 {population} 条完整三元组。
- 样本量：{sample_size} 条。
- 随机种子：{seed}。
- 抽样方式：固定种子的简单随机抽样，不放回。在总体为 {population}、样本量为 {sample_size}、95% 置信水平和最保守比例 0.5 时，有限总体误差界限约为 ±{worst_case_margin:.1%}；实际结果同时报告 Wilson 95% 置信区间。该口径可以估计总体三元组精确率，但当前数据缺少来源文档和设施标签，无法预先按 BSRF/HEPS 或文档来源分层。

## 独立评审

两位领域专家分别填写 `expert_1_annotations.csv` 和 `expert_2_annotations.csv`，在首次评审完成前不交换判断。每条三元组必须先补充 `facility_BSRF_or_HEPS`、`source_document`、`source_location` 和 `source_excerpt`。没有可核对来源时，将 `source_support_Y_N_UNCERTAIN` 和 `triple_correct_Y_N_UNCERTAIN` 标记为 `UNCERTAIN`，不要推测为正确。

`triple_correct_Y_N_UNCERTAIN` 采用严格判定：只有主题、关系、客体、实体类型均正确且有来源支持时才填 `Y`；任何核心组成错误均填 `N`。`action_KEEP_CORRECT_REMOVE_UNCERTAIN` 只允许 `KEEP`、`CORRECT`、`REMOVE` 或 `UNCERTAIN`。若选择 `CORRECT`，应填写相应的 corrected 字段。

错误类型使用分号分隔的受控代码：{error_codes}。

## 一致性与裁决

两位专家完成后，运行 `calculate_kg_validation_metrics.py`。脚本将计算严格三元组精确率及 Wilson 95% 置信区间、各组成项正确率、原始一致率和 Cohen's kappa，并输出完整裁决模板及分歧清单。第三位专家或两位专家共同裁决后，填写 `adjudication_template.csv` 并通过 `--adjudicated` 传入脚本。裁决表中的处理数量只代表抽样样本；全图谱实际纠正、删除、同义词合并和关系规范化数量必须登记在 `kg_cleaning_log_template.csv` 中，再通过 `--cleaning-log` 汇总。

## Recall 的限制

从已抽取三元组中抽样只能估计 precision，不能估计 recall。Recall 需要独立随机抽取来源文档中的段落或其他预先定义的来源单元，由专家在不查看系统结果的情况下穷举金标准三元组，再计算系统匹配到的金标准三元组比例。当前三元组文件没有段落级 provenance，因此不能仅凭该文件给出可信 recall。
""",
        encoding="utf-8",
    )


def write_recall_templates(output_dir: Path) -> None:
    source_unit_columns = [
        "source_unit_id",
        "included_in_random_sample_Y_N",
        "facility_BSRF_or_HEPS",
        "source_document",
        "source_location",
        "source_text",
        "expert_1_completed_Y_N",
        "expert_2_completed_Y_N",
        "adjudicated_Y_N",
        "notes",
    ]
    with (output_dir / "recall_source_units.csv").open(
        "w", encoding="utf-8-sig", newline=""
    ) as handle:
        csv.writer(handle).writerow(source_unit_columns)

    columns = [
        "source_unit_id",
        "gold_triple_id",
        "gold_sub_name",
        "gold_sub_type",
        "gold_rel_name",
        "gold_rel_type",
        "gold_obj_name",
        "gold_obj_type",
        "matched_by_system_Y_N_UNCERTAIN",
        "matched_system_sub_name",
        "matched_system_rel_name",
        "matched_system_obj_name",
        "notes",
    ]
    with (output_dir / "recall_gold_triples.csv").open(
        "w", encoding="utf-8-sig", newline=""
    ) as handle:
        csv.writer(handle).writerow(columns)


def main() -> None:
    args = parse_args()
    if args.sample_size <= 0:
        raise ValueError("sample-size 必须为正整数")

    source_file = args.input.resolve()
    population = load_unique_triples(source_file, args.sheet)
    if args.sample_size > len(population):
        raise ValueError(
            f"样本量 {args.sample_size} 大于唯一完整三元组总数 {len(population)}"
        )

    rng = random.Random(args.seed)
    selected_indexes = sorted(rng.sample(range(len(population)), args.sample_size))
    sampled = [population[index] for index in selected_indexes]
    annotation_rows = make_annotation_rows(sampled, source_file, args.sheet)

    output_dir = args.output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    write_csv(output_dir / "expert_1_annotations.csv", annotation_rows)
    write_csv(output_dir / "expert_2_annotations.csv", annotation_rows)
    write_protocol(
        output_dir / "annotation_protocol.md",
        population=len(population),
        sample_size=args.sample_size,
        seed=args.seed,
    )
    write_recall_templates(output_dir)
    print(f"验证材料已生成：{output_dir}")


if __name__ == "__main__":
    main()
