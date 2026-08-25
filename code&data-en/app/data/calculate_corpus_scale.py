#!/usr/bin/env python3
"""CalculateBSRF-HEPS K-V Corpus size.

By default, the files in the same directory as this script are read.BSRF-HEPS-KnowledgeBase-k-v.xlsx, and output in terminal
JSON Format statistics results. When you need to save the results, you can use--json-output SpecifyJSON File path.
"""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from statistics import mean, median
from typing import Any

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = SCRIPT_DIR / "BSRF-HEPS-KnowledgeBase-k-v.xlsx"
DEFAULT_SHEET = "Sheet1"
DEFAULT_PLACEHOLDERS = ("[Picture placeholder]", "——")
WHITESPACE_RE = re.compile(r"\s+")
ENGLISH_WORD_RE = re.compile(r"[A-Za-z]+(?:[-'][A-Za-z]+)*")
CHINESE_CHAR_RE = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff]")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="CalculateExcel K-V Number of records, number of deduplications and text size of the corpus."
    )
    parser.add_argument(
        "input_file",
        nargs="?",
        type=Path,
        default=DEFAULT_INPUT,
        help=f"inputExcel File, default is:{DEFAULT_INPUT.name}",
    )
    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET,
        help=f"The worksheet to be counted, the default is:{DEFAULT_SHEET}",
    )
    parser.add_argument(
        "--json-output",
        type=Path,
        help="Optional, save statistical results asJSON File",
    )
    parser.add_argument(
        "--include-placeholders",
        action="store_true",
        help="will[Picture placeholder] and - deemed to be substantial textual records",
    )
    return parser.parse_args()


def as_text(value: Any) -> str:
    """willExcel The cells are converted into text that can be counted, and the leading and trailing blanks are removed."""
    return "" if value is None else str(value).strip()


def length_stats(values: list[str]) -> dict[str, int | float]:
    lengths = [len(value) for value in values]
    if not lengths:
        return {"total": 0, "mean": 0, "median": 0, "min": 0, "max": 0}
    return {
        "total": sum(lengths),
        "mean": round(mean(lengths), 2),
        "median": median(lengths),
        "min": min(lengths),
        "max": max(lengths),
    }


def count_without_whitespace(values: list[str]) -> int:
    return sum(len(WHITESPACE_RE.sub("", value)) for value in values)


def count_chinese_characters(values: list[str]) -> int:
    return sum(len(CHINESE_CHAR_RE.findall(value)) for value in values)


def count_english_words(values: list[str]) -> int:
    return sum(len(ENGLISH_WORD_RE.findall(value)) for value in values)


def duplicate_summary(values: list[Any]) -> dict[str, int]:
    counts = Counter(values)
    duplicate_groups = {value: count for value, count in counts.items() if count > 1}
    return {
        "duplicate_groups": len(duplicate_groups),
        "rows_in_duplicate_groups": sum(duplicate_groups.values()),
        "extra_duplicate_rows": sum(count - 1 for count in duplicate_groups.values()),
    }


def try_count_tokens(values: list[str]) -> dict[str, Any]:
    """If installedtiktoken, then additional statisticscl100k_base token number; otherwise, the uninstalled status is returned.

    cl100k_base Noall-MiniLM-L6-v2 of nativetokenizer, Therefore, this indicator is only used for
    A reproducible approximate scale reference that should not pass off as a true modeltoken Count.
    """
    try:
        import tiktoken  # type: ignore
    except ImportError:
        return {"available": False, "encoding": None, "total": None}

    encoding_name = "cl100k_base"
    encoder = tiktoken.get_encoding(encoding_name)
    return {
        "available": True,
        "encoding": encoding_name,
        "total": sum(len(encoder.encode(value)) for value in values),
    }


def inspect_embedding_file(file_path: Path) -> dict[str, Any]:
    """Read existing vectorJSON, Check the number of statements, number of vectors and vector dimensions."""
    if not file_path.is_file():
        return {"available": False, "file": str(file_path)}

    try:
        payload = json.loads(file_path.read_text(encoding="utf-8"))
        sentences = payload.get("sentences", [])
        embeddings = payload.get("embeddings", [])
        dimensions = sorted(
            {len(embedding) for embedding in embeddings if isinstance(embedding, list)}
        )
        return {
            "available": True,
            "file": str(file_path.resolve()),
            "sentence_count": len(sentences),
            "embedding_count": len(embeddings),
            "dimensions": dimensions,
        }
    except (OSError, UnicodeError, json.JSONDecodeError, AttributeError) as exc:
        return {
            "available": False,
            "file": str(file_path.resolve()),
            "error": str(exc),
        }


def load_kv_records(input_file: Path, sheet_name: str) -> tuple[int, list[dict[str, Any]]]:
    if not input_file.is_file():
        raise FileNotFoundError(f"Input file not found:{input_file}")

    workbook = load_workbook(input_file, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"worksheet{sheet_name!r} Does not exist, available worksheets:{workbook.sheetnames}"
            )
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ValueError(f"worksheet{sheet_name!r} is empty")

        headers = [as_text(value) for value in header]
        try:
            key_index = headers.index("k")
            value_index = headers.index("v")
        except ValueError as exc:
            raise ValueError(
                f"The first row of the worksheet must containk andv Column, the actual header is:{headers}"
            ) from exc

        records = []
        for excel_row, row in enumerate(rows, start=2):
            key = as_text(row[key_index] if key_index < len(row) else None)
            value = as_text(row[value_index] if value_index < len(row) else None)
            if key or value:
                records.append(
                    {"excel_row": excel_row, "key": key, "value": value}
                )
        return worksheet.max_row, records
    finally:
        workbook.close()


def build_report(
    input_file: Path,
    sheet_name: str,
    include_placeholders: bool,
) -> dict[str, Any]:
    physical_max_row, records = load_kv_records(input_file, sheet_name)
    placeholders = set() if include_placeholders else set(DEFAULT_PLACEHOLDERS)

    missing_key_rows = [record for record in records if not record["key"]]
    missing_value_rows = [record for record in records if not record["value"]]
    complete_records = [
        record for record in records if record["key"] and record["value"]
    ]
    substantive_records = [
        record
        for record in complete_records
        if record["value"] not in placeholders
    ]

    keys = [record["key"] for record in complete_records]
    values = [record["value"] for record in complete_records]
    pairs = [(record["key"], record["value"]) for record in complete_records]
    substantive_keys = [record["key"] for record in substantive_records]
    substantive_values = [record["value"] for record in substantive_records]
    substantive_pairs = [
        (record["key"], record["value"]) for record in substantive_records
    ]

    values_by_key: defaultdict[str, set[str]] = defaultdict(set)
    rows_by_key: defaultdict[str, list[int]] = defaultdict(list)
    for record in complete_records:
        values_by_key[record["key"]].add(record["value"])
        rows_by_key[record["key"]].append(record["excel_row"])

    conflicting_keys = {
        key: {
            "row_numbers": row_numbers,
            "value_count": len(values_by_key[key]),
        }
        for key, row_numbers in rows_by_key.items()
        if len(values_by_key[key]) > 1
    }

    combined_text = [key + value for key, value in pairs]
    substantive_combined_text = [key + value for key, value in substantive_pairs]
    placeholder_counts = Counter(
        record["value"]
        for record in complete_records
        if record["value"] in DEFAULT_PLACEHOLDERS
    )

    report: dict[str, Any] = {
        "input_file": str(input_file.resolve()),
        "sheet": sheet_name,
        "physical_max_row": physical_max_row,
        "records": {
            "raw_nonempty_kv_rows": len(records),
            "complete_kv_rows": len(complete_records),
            "substantive_kv_rows": len(substantive_records),
            "missing_key_rows": len(missing_key_rows),
            "missing_value_rows": len(missing_value_rows),
        },
        "uniqueness": {
            "unique_keys": len(set(keys)),
            "unique_values": len(set(values)),
            "unique_kv_pairs": len(set(pairs)),
            "substantive_unique_keys": len(set(substantive_keys)),
            "substantive_unique_values": len(set(substantive_values)),
            "substantive_unique_kv_pairs": len(set(substantive_pairs)),
            "duplicate_keys": duplicate_summary(keys),
            "duplicate_values": duplicate_summary(values),
            "duplicate_kv_pairs": duplicate_summary(pairs),
            "conflicting_key_groups": len(conflicting_keys),
            "conflicting_keys": conflicting_keys,
        },
        "placeholders": {
            "excluded_from_substantive_count": not include_placeholders,
            "markers": list(DEFAULT_PLACEHOLDERS),
            "counts": dict(placeholder_counts),
            "total": sum(placeholder_counts.values()),
        },
        "text_size": {
            "key_characters": length_stats(keys),
            "value_characters": length_stats(values),
            "key_value_characters": length_stats(combined_text),
            "key_characters_without_whitespace": count_without_whitespace(keys),
            "value_characters_without_whitespace": count_without_whitespace(values),
            "key_value_characters_without_whitespace": count_without_whitespace(
                combined_text
            ),
            "key_chinese_characters": count_chinese_characters(keys),
            "value_chinese_characters": count_chinese_characters(values),
            "key_value_chinese_characters": count_chinese_characters(combined_text),
            "key_english_word_units": count_english_words(keys),
            "value_english_word_units": count_english_words(values),
            "substantive_key_value_characters": length_stats(
                substantive_combined_text
            ),
            "substantive_key_value_characters_without_whitespace": count_without_whitespace(
                substantive_combined_text
            ),
            "approximate_tokens": {
                "all_records": try_count_tokens(combined_text),
                "substantive_records": try_count_tokens(substantive_combined_text),
                "note": (
                    "Token Number is only installedtiktoken is calculated and usedcl100k_base; "
                    "This value is notall-MiniLM-L6-v2 of nativetoken Count."
                ),
            },
        },
        "vector_index": {
            "expected_records_from_excel": len(records),
            "model": "all-MiniLM-L6-v2",
            "key_embeddings": inspect_embedding_file(
                input_file.parent / "k_embeddings.json"
            ),
            "value_embeddings": inspect_embedding_file(
                input_file.parent / "v_embeddings.json"
            ),
        },
    }
    return report


def main() -> None:
    args = parse_args()
    report = build_report(args.input_file.resolve(), args.sheet, args.include_placeholders)
    output = json.dumps(report, ensure_ascii=False, indent=2)
    print(output)

    if args.json_output:
        json_path = args.json_output.resolve()
        json_path.parent.mkdir(parents=True, exist_ok=True)
        json_path.write_text(output + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
